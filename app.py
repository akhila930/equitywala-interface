from io import BytesIO
import xlsxwriter
from datetime import datetime, timedelta
from datetime import datetime, timedelta, date
from flask import jsonify, request
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file, make_response
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import os
from werkzeug.utils import secure_filename
import pandas as pd
import json
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import pdfkit
from num2words import num2words
from fpdf import FPDF
from openpyxl import Workbook
import tempfile
import xlsxwriter
import numpy as np
from urllib.parse import urlparse
import random
import string
from sqlalchemy import extract, func
import time
from io import BytesIO
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill



app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///users.db'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xlsx', 'xls', 'jpg', 'jpeg', 'png'}
db = SQLAlchemy(app)

# Configure upload folder
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Configure allowed file extensions
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg'}

def allowed_file(filename):
    # Allow all file types
    return True

# Custom filter for currency formatting
@app.template_filter('format_currency')
def format_currency(value):
    if value is None:
        return "₹0.00"
    return f"₹{value:,.2f}"

# Create uploads directory if it doesn't exist
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# Add this constant at the top of the file, after the imports
MASTER_ACCESS_CODE = "EQ#2024@SEC$KEY"  # Complex master key for confidential access

# Models
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    is_manager = db.Column(db.Boolean, default=False)
    department = db.Column(db.String(100))  # Add department
    hierarchy_level = db.Column(db.Integer, default=0)  # 0: Employee, 1: Manager, 2: Head, 3: Admin

class Employee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True)
    phone = db.Column(db.String(20))
    dob = db.Column(db.Date)
    role = db.Column(db.String(100))
    department = db.Column(db.String(100))
    employee_type = db.Column(db.String(50))  # full-time, part-time, contract, intern
    salary = db.Column(db.Float)
    pay_frequency = db.Column(db.String(20), default='monthly')  # monthly, bi-weekly, weekly
    join_date = db.Column(db.DateTime, default=datetime.utcnow)
    profile_photo = db.Column(db.String(500))  # Path to profile photo
    
    # Relationships with unique backref names
    tasks = db.relationship('Task', backref='task_employee', lazy=True)
    attendances = db.relationship('Attendance', backref='attendance_employee', lazy=True)
    milestones = db.relationship('Milestone', backref='milestone_employee', lazy=True)
    documents = db.relationship('Document', backref='document_employee', lazy=True)
    salary_components = db.relationship('SalaryComponent', backref='salary_component_employee', lazy=True)
    salaries = db.relationship('EmployeeSalary', backref='salary_employee', lazy=True)
    assigned_leads = db.relationship('Lead', backref='lead_employee', lazy=True)
    assigned_clients = db.relationship('Client', backref='client_employee', lazy=True)
    received_feedback = db.relationship('EmployeeFeedback', back_populates='employee', lazy=True)

    def __repr__(self):
        return f'<Employee {self.name}>'

class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(20), nullable=False)  # present, absent, half-day
    comments = db.Column(db.Text)

    def __repr__(self):
        return f'<Attendance {self.employee_id} {self.date} {self.status}>'

class Task(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    due_date = db.Column(db.DateTime)
    priority = db.Column(db.String(50), default='medium')  # high, medium, low
    status = db.Column(db.String(50), default='pending')  # pending, in-progress, completed
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'))
    assigned_by_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_date = db.Column(db.DateTime, default=datetime.utcnow)

    assigned_by = db.relationship('User', backref='assigned_tasks')

class Candidate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120))
    phone = db.Column(db.String(20))
    role_applied = db.Column(db.String(100))
    experience = db.Column(db.Float)
    source = db.Column(db.String(50))  # LinkedIn, Job Portal, Referral, etc.
    status = db.Column(db.String(50))  # applied, screening, interview, selected, rejected
    resume_path = db.Column(db.String(500))
    resume_link = db.Column(db.String(500))
    created_date = db.Column(db.DateTime, default=datetime.utcnow)
    notes = db.Column(db.Text)

class Interview(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    candidate_id = db.Column(db.Integer, db.ForeignKey('candidate.id'))
    scheduled_date = db.Column(db.DateTime)
    interviewer_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    status = db.Column(db.String(50))  # scheduled, completed, cancelled
    feedback = db.Column(db.Text)
    rating = db.Column(db.Integer)

class SocialMediaCampaign(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    budget = db.Column(db.Float, default=0.0)
    amount_spent = db.Column(db.Float, default=0.0)
    start_date = db.Column(db.DateTime)
    end_date = db.Column(db.DateTime)
    status = db.Column(db.String(50))
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class ScheduledPost(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    content = db.Column(db.Text, nullable=False)
    platforms = db.Column(db.String(200))  # Comma-separated list of platforms
    scheduled_time = db.Column(db.DateTime)
    status = db.Column(db.String(50))
    campaign_id = db.Column(db.Integer, db.ForeignKey('social_media_campaign.id'))
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Department(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    budget = db.Column(db.Float, default=0.0)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    expenses = db.relationship('Expense', backref='department', lazy=True)

class Expense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.String(200), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    date = db.Column(db.DateTime, nullable=False)
    receipt_path = db.Column(db.String(500))
    department_id = db.Column(db.Integer, db.ForeignKey('department.id'))
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class DocumentCategory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    is_confidential = db.Column(db.Boolean, default=False)
    access_code = db.Column(db.String(100))  # For confidential categories

class Document(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    document_type = db.Column(db.String(50))  # personal or company
    category_id = db.Column(db.Integer, db.ForeignKey('document_category.id'))
    description = db.Column(db.Text)
    uploaded_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    uploaded_date = db.Column(db.DateTime, default=datetime.utcnow)
    version = db.Column(db.Integer, default=1)
    is_latest = db.Column(db.Boolean, default=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'))
    
    category = db.relationship('DocumentCategory', backref='documents')
    uploader = db.relationship('User', backref='uploaded_documents')
    # Simple relationship without backref since it's defined in Employee model
    employee = db.relationship('Employee')

class DocumentAccessLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    document_id = db.Column(db.Integer, db.ForeignKey('document.id'))
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    action = db.Column(db.String(50))  # upload, download, view
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    
    document = db.relationship('Document')
    user = db.relationship('User')

class Milestone(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    date = db.Column(db.DateTime, nullable=False)
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'))
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class EmployeeFeedback(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=False)
    reviewer_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    feedback_type = db.Column(db.String(50), nullable=False)  # manager, peer, self
    rating = db.Column(db.Integer, nullable=False)  # 1-5
    comments = db.Column(db.Text)
    created_date = db.Column(db.DateTime, default=datetime.utcnow)
    updated_date = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    employee = db.relationship('Employee', back_populates='received_feedback')
    reviewer = db.relationship('User', backref='given_feedback')

class ClientFeedback(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_name = db.Column(db.String(100))
    client_email = db.Column(db.String(120))
    rating = db.Column(db.Integer)  # Changed to allow null values initially
    comments = db.Column(db.Text)
    feedback_type = db.Column(db.String(50), default='client')
    created_date = db.Column(db.DateTime, default=datetime.utcnow)
    feedback_link_id = db.Column(db.String(50), unique=True)  # Unique identifier for feedback link

class FeedbackAuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    feedback_type = db.Column(db.String(50), nullable=False)  # employee or client
    feedback_id = db.Column(db.Integer, nullable=False)  # ID of the feedback
    action = db.Column(db.String(50), nullable=False)  # create, update, delete
    performed_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    action_date = db.Column(db.DateTime, default=datetime.utcnow)
    details = db.Column(db.Text)  # Additional details about the action

class LeaveType(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)  # e.g., Annual, Sick, Casual
    default_days = db.Column(db.Integer, default=0)  # Default number of days per year
    color_code = db.Column(db.String(7))  # Hex color code for calendar display

class LeaveRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    leave_type_id = db.Column(db.Integer, db.ForeignKey('leave_type.id'), nullable=False)
    start_date = db.Column(db.DateTime, nullable=False)
    end_date = db.Column(db.DateTime, nullable=False)
    reason = db.Column(db.Text)
    status = db.Column(db.String(20), default='pending')  # pending, approved, rejected
    approver_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    employee = db.relationship('User', foreign_keys=[employee_id], backref='leave_requests')
    leave_type = db.relationship('LeaveType')
    approver = db.relationship('User', foreign_keys=[approver_id], backref='leave_approvals')

class LeaveComment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    leave_request_id = db.Column(db.Integer, db.ForeignKey('leave_request.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    comment = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    leave_request = db.relationship('LeaveRequest', backref='comments')
    user = db.relationship('User')

class Lead(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120))
    phone = db.Column(db.String(20))
    source = db.Column(db.String(50))  # Website, Referral, Social Media, etc.
    status = db.Column(db.String(50), default='new')  # new, contacted, negotiating, closed
    score = db.Column(db.Integer, default=0)  # Lead score 0-100
    assigned_to = db.Column(db.Integer, db.ForeignKey('employee.id'))
    notes = db.Column(db.Text)
    created_date = db.Column(db.DateTime, default=datetime.utcnow)
    updated_date = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Simple relationship without backref since it's defined in Employee model
    employee = db.relationship('Employee')

class Client(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120))
    phone = db.Column(db.String(20))
    company = db.Column(db.String(100))
    address = db.Column(db.Text)
    status = db.Column(db.String(50), default='active')  # active, inactive
    assigned_to = db.Column(db.Integer, db.ForeignKey('employee.id'))
    created_date = db.Column(db.DateTime, default=datetime.utcnow)
    updated_date = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Simple relationship without backref since it's defined in Employee model
    employee = db.relationship('Employee')
    documents = db.relationship('ClientDocument', backref='client', cascade='all, delete-orphan')
    interactions = db.relationship('ClientInteraction', backref='client', cascade='all, delete-orphan')
    services = db.relationship('ClientService', backref='client', cascade='all, delete-orphan')

class ClientDocument(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('client.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    file_path = db.Column(db.String(255), nullable=False)
    document_type = db.Column(db.String(50))  # contract, invoice, proposal, etc.
    upload_date = db.Column(db.DateTime, default=datetime.utcnow)

class ClientInteraction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('client.id'), nullable=False)
    interaction_type = db.Column(db.String(50))  # meeting, call, email, etc.
    summary = db.Column(db.Text)
    interaction_date = db.Column(db.DateTime, default=datetime.utcnow)
    next_followup_date = db.Column(db.DateTime)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))

    user = db.relationship('User')

class ClientService(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('client.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    status = db.Column(db.String(50), default='pending')  # pending, in_progress, completed, cancelled
    created_date = db.Column(db.DateTime, default=datetime.utcnow)
    updated_date = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Service-specific documents and interactions
    documents = db.relationship('ServiceDocument', backref='service', cascade='all, delete-orphan')
    interactions = db.relationship('ServiceInteraction', backref='service', cascade='all, delete-orphan')

class ServiceDocument(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    service_id = db.Column(db.Integer, db.ForeignKey('client_service.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    file_path = db.Column(db.String(255), nullable=False)
    document_type = db.Column(db.String(50))  # contract, invoice, proposal, etc.
    upload_date = db.Column(db.DateTime, default=datetime.utcnow)

class ServiceInteraction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    service_id = db.Column(db.Integer, db.ForeignKey('client_service.id'), nullable=False)
    interaction_type = db.Column(db.String(50))  # meeting, call, email, etc.
    summary = db.Column(db.Text)
    interaction_date = db.Column(db.DateTime, default=datetime.utcnow)
    next_followup_date = db.Column(db.DateTime)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))

    user = db.relationship('User')

class SalaryComponent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    type = db.Column(db.String(50))  # allowance, deduction
    is_percentage = db.Column(db.Boolean, default=False)
    value = db.Column(db.Float)  # Fixed amount or percentage value
    is_taxable = db.Column(db.Boolean, default=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Simple relationship without backref since it's defined in Employee model
    employee = db.relationship('Employee')

class EmployeeSalary(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=False)
    basic_pay = db.Column(db.Float, nullable=False)
    month = db.Column(db.Integer, nullable=False)
    year = db.Column(db.Integer, nullable=False)
    status = db.Column(db.String(50), default='pending')  # pending, processed
    net_salary = db.Column(db.Float)
    generated_date = db.Column(db.DateTime, default=datetime.utcnow)
    processed_date = db.Column(db.DateTime)
    processed_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    
    # Simple relationship without backref since it's defined in Employee model
    employee = db.relationship('Employee')
    processor = db.relationship('User', backref='processed_salaries')

class SalaryDetail(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    salary_id = db.Column(db.Integer, db.ForeignKey('employee_salary.id'), nullable=False)
    component_id = db.Column(db.Integer, db.ForeignKey('salary_component.id'), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    
    salary = db.relationship('EmployeeSalary', backref='details')
    component = db.relationship('SalaryComponent')

# Routes
@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        
        user = User.query.filter_by(email=email).first()
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            session['user_name'] = user.name
            return redirect(url_for('dashboard'))
        
        flash('Invalid email or password')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('dashboard.html', user_name=session.get('user_name'))

@app.route('/employee-dashboard')
def employee_dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    employees = Employee.query.all()
    return render_template('employee_dashboard.html', 
                         employees=employees,
                         user_name=session.get('user_name'))

@app.route('/employee-profile/<int:employee_id>')
def employee_profile(employee_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    employee = Employee.query.get_or_404(employee_id)
    productivity_score = calculate_productivity_score(employee)
    attendance_percentage = calculate_attendance_percentage(employee)
    
    # Get all tasks for the employee
    tasks = Task.query.filter_by(employee_id=employee_id).order_by(Task.due_date.desc()).all()
    
    # Get milestones
    milestones = Milestone.query.filter_by(employee_id=employee_id).order_by(Milestone.date.desc()).all()
    
    # Get documents separated by type
    personal_docs = Document.query.filter_by(
        employee_id=employee_id,
        document_type='personal'
    ).order_by(Document.uploaded_date.desc()).all()
    
    company_docs = Document.query.filter_by(
        employee_id=employee_id,
        document_type='company'
    ).order_by(Document.uploaded_date.desc()).all()

    # Get document categories for the upload form
    document_categories = DocumentCategory.query.all()

    # Get attendance data for the current month
    today = datetime.now()
    current_month_attendance = Attendance.query.filter(
        Attendance.employee_id == employee_id,
        extract('month', Attendance.date) == today.month,
        extract('year', Attendance.date) == today.year
    ).all()

    # Get current month's attendance
    start_date = datetime(today.year, today.month, 1)
    if today.month == 12:
        end_date = datetime(today.year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = datetime(today.year, today.month + 1, 1) - timedelta(days=1)
    
    attendance_records = Attendance.query.filter(
        Attendance.employee_id == employee_id,
        Attendance.date >= start_date,
        Attendance.date <= end_date
    ).all()
    
    # Create a calendar with attendance status
    calendar_days = []
    current_date = start_date
    while current_date <= end_date:
        attendance = next(
            (a for a in attendance_records if a.date == current_date.date()),
            None
        )
        calendar_days.append({
            'date': current_date,
            'status': attendance.status if attendance else 'unmarked'
        })
        current_date += timedelta(days=1)
    
    return render_template('employee_profile.html',
                         employee=employee,
                         productivity_score=productivity_score,
                         attendance_percentage=attendance_percentage,
                         tasks=tasks,
                         milestones=milestones,
                         personal_docs=personal_docs,
                         company_docs=company_docs,
                         document_categories=document_categories,
                         current_month_attendance=current_month_attendance,
                         calendar_days=calendar_days,
                         current_month=start_date.strftime('%B %Y'),
                         user_name=session.get('user_name'),
                         datetime=datetime)

def calculate_productivity_score(employee):
    # Implement your productivity score calculation logic
    # This is a placeholder implementation
    completed_tasks = Task.query.filter_by(
        employee_id=employee.id,
        status='completed'
    ).count()
    return completed_tasks * 10

def calculate_attendance_percentage(employee):
    # Calculate attendance percentage for the current month
    today = datetime.now()
    start_date = date(today.year, today.month, 1)
    if today.month == 12:
        end_date = date(today.year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(today.year, today.month + 1, 1) - timedelta(days=1)
    
    working_days = np.busday_count(start_date, end_date)
    if working_days == 0:
        return 0.0
        
    present_days = Attendance.query.filter(
        Attendance.employee_id == employee.id,
        Attendance.date >= start_date,
        Attendance.date <= end_date,
        Attendance.status.in_(['present', 'half-day'])
    ).count()
    
    return round((present_days / working_days) * 100, 2)

@app.route('/recruitment-dashboard')
def recruitment_dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    candidates = Candidate.query.order_by(Candidate.created_date.desc()).all()
    interviews = Interview.query.filter(
        Interview.scheduled_date >= datetime.now()
    ).order_by(Interview.scheduled_date).all()
    
    return render_template('recruitment_dashboard.html',
                         candidates=candidates,
                         interviews=interviews,
                         user_name=session.get('user_name'))

@app.route('/add-candidate', methods=['GET', 'POST'])
def add_candidate():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        try:
            # Handle file upload
            resume = request.files.get('resume')
            resume_path = None
            if resume and allowed_file(resume.filename):
                filename = secure_filename(f"{int(time.time())}_{resume.filename}")
                resume_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                resume.save(resume_path)
            
            # Create new candidate
            candidate = Candidate(
                name=request.form.get('name'),
                email=request.form.get('email'),
                phone=request.form.get('phone'),
                role_applied=request.form.get('position'),  # Changed from 'position' to 'role_applied'
                experience=float(request.form.get('experience', 0)),  # Add default value
                status=request.form.get('status', 'Applied'),  # Add default value
                resume_path=resume_path,
                created_date=datetime.utcnow()
            )
            
            db.session.add(candidate)
            db.session.commit()
            
            flash('Candidate added successfully')
            return redirect(url_for('recruitment_dashboard'))
        except Exception as e:
            flash(f'Error adding candidate: {str(e)}')
    
    return render_template('add_candidate.html', user_name=session.get('user_name'))

@app.route('/update-candidate-status/<int:candidate_id>', methods=['POST'])
def update_candidate_status(candidate_id):
    candidate = Candidate.query.get_or_404(candidate_id)
    try:
        candidate.status = request.form.get('status')
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/schedule-interview/<int:candidate_id>', methods=['POST'])
def schedule_interview(candidate_id):
    try:
        interview = Interview(
            candidate_id=candidate_id,
            scheduled_date=datetime.strptime(request.form.get('scheduled_date'), '%Y-%m-%dT%H:%M'),
            interviewer_id=session['user_id'],
            status='scheduled'
        )
        db.session.add(interview)
        db.session.commit()
        flash('Interview scheduled successfully')
    except Exception as e:
        flash(f'Error scheduling interview: {str(e)}')
    
    return redirect(url_for('recruitment_dashboard'))

def allowed_file(filename):
    # Allow all file types
    return True

@app.route('/social-media-dashboard')
def social_media_dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get all campaigns and filter active ones
    campaigns = SocialMediaCampaign.query.all()
    active_campaigns = SocialMediaCampaign.query.filter_by(status='active').all()
    
    # Get scheduled posts ordered by scheduled time
    scheduled_posts = ScheduledPost.query.order_by(ScheduledPost.scheduled_time).all()
    
    # Add campaign names to scheduled posts for display
    for post in scheduled_posts:
        if post.campaign_id:
            campaign = next((c for c in campaigns if c.id == post.campaign_id), None)
            post.campaign_name = campaign.name if campaign else "N/A"
        else:
            post.campaign_name = "N/A"
    
    return render_template('social_media_dashboard.html',
                         campaigns=campaigns,
                         active_campaigns=active_campaigns,
                         scheduled_posts=scheduled_posts,
                         user_name=session.get('user_name'))

@app.route('/add-campaign', methods=['POST'])
def add_campaign():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    try:
        campaign = SocialMediaCampaign(
            name=request.form.get('name'),
            budget=float(request.form.get('budget')),
            amount_spent=float(request.form.get('amount_spent', 0)),
            start_date=datetime.strptime(request.form.get('start_date'), '%Y-%m-%d'),
            end_date=datetime.strptime(request.form.get('end_date'), '%Y-%m-%d'),
            status='Active',
            created_by=session['user_id']
        )
        db.session.add(campaign)
        db.session.commit()
        flash('Campaign created successfully')
    except Exception as e:
        flash(f'Error creating campaign: {str(e)}')
    
    return redirect(url_for('social_media_dashboard'))

@app.route('/add-post', methods=['POST'])
def add_post():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    try:
        platforms = ','.join(request.form.getlist('platforms'))
        post = ScheduledPost(
            content=request.form.get('content'),
            platforms=platforms,
            scheduled_time=datetime.strptime(request.form.get('scheduled_time'), '%Y-%m-%dT%H:%M'),
            status='Scheduled',
            campaign_id=request.form.get('campaign_id'),
            created_by=session['user_id']
        )
        db.session.add(post)
        db.session.commit()
        flash('Post scheduled successfully')
    except Exception as e:
        flash(f'Error scheduling post: {str(e)}')
    
    return redirect(url_for('social_media_dashboard'))

@app.route('/budget-dashboard')
def budget_dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    departments = Department.query.all()
    expenses = Expense.query.order_by(Expense.date.desc()).all()
    
    total_budget = sum(dept.budget for dept in departments)
    total_spent = sum(expense.amount for expense in expenses)
    
    # Calculate monthly expenses for the chart
    monthly_expenses = {}
    for expense in expenses:
        month_key = expense.date.strftime('%Y-%m')
        monthly_expenses[month_key] = monthly_expenses.get(month_key, 0) + expense.amount
    
    return render_template('budget_dashboard.html',
                         departments=departments,
                         expenses=expenses,
                         total_budget=total_budget,
                         total_spent=total_spent,
                         monthly_expenses=monthly_expenses,
                         user_name=session.get('user_name'))

@app.route('/add-department', methods=['POST'])
def add_department():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    try:
        department = Department(
            name=request.form.get('name'),
            budget=float(request.form.get('budget')),
            created_by=session['user_id']
        )
        db.session.add(department)
        db.session.commit()
        flash('Department added successfully')
    except Exception as e:
        flash(f'Error adding department: {str(e)}')
    
    return redirect(url_for('budget_dashboard'))

@app.route('/add-expense', methods=['POST'])
def add_expense():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    try:
        receipt = request.files.get('receipt')
        receipt_path = None
        
        if receipt and allowed_file(receipt.filename):
            # Create receipts directory if it doesn't exist
            receipts_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'receipts')
            if not os.path.exists(receipts_dir):
                os.makedirs(receipts_dir)
            
            # Save receipt with timestamp to avoid filename conflicts
            filename = secure_filename(f"{int(time.time())}_{receipt.filename}")
            receipt_path = os.path.join('receipts', filename)  # Store relative path
            full_path = os.path.join(app.config['UPLOAD_FOLDER'], receipt_path)
            receipt.save(full_path)
        
        expense = Expense(
            description=request.form.get('description'),
            amount=float(request.form.get('amount')),
            date=datetime.strptime(request.form.get('date'), '%Y-%m-%d'),
            department_id=request.form.get('department_id'),
            receipt_path=receipt_path,
            created_by=session['user_id']
        )
        db.session.add(expense)
        db.session.commit()
        flash('Expense added successfully')
    except Exception as e:
        flash(f'Error adding expense: {str(e)}')
    
    return redirect(url_for('budget_dashboard'))

@app.route('/add-employee', methods=['GET', 'POST'])
def add_employee():
    if request.method == 'POST':
        try:
            # Get form data
            name = request.form['name']
            email = request.form['email']
            phone = request.form['phone']
            role = request.form['role']
            department = request.form['department']
            employee_type = request.form['employee_type']
            salary = float(request.form['salary'])
            join_date = datetime.strptime(request.form['join_date'], '%Y-%m-%d')
            
            # Create new employee
            employee = Employee(
                name=name,
                email=email,
                phone=phone,
                role=role,
                department=department,
                employee_type=employee_type,
                salary=salary,
                join_date=join_date
            )
            
            db.session.add(employee)
            db.session.commit()
            
            flash('Employee added successfully', 'success')
            return redirect(url_for('employee_dashboard'))
            
        except Exception as e:
            db.session.rollback()
            flash('Error adding employee. Please try again.', 'error')
            return redirect(url_for('add_employee'))
    
    return render_template('add_employee.html')

@app.route('/edit-employee/<int:employee_id>', methods=['GET', 'POST'])
def edit_employee(employee_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    employee = Employee.query.get_or_404(employee_id)
    
    if request.method == 'POST':
        try:
            employee.name = request.form.get('name')
            employee.email = request.form.get('email')
            employee.phone = request.form.get('phone')
            employee.dob = datetime.strptime(request.form.get('dob'), '%Y-%m-%d').date()
            employee.role = request.form.get('role')
            employee.department = request.form.get('department')
            employee.employee_type = request.form.get('employee_type')
            employee.salary = float(request.form.get('salary'))
            employee.pay_frequency = request.form.get('pay_frequency')
            employee.join_date = datetime.strptime(request.form.get('join_date'), '%Y-%m-%d')
            
            # Handle profile photo upload
            profile_photo = request.files.get('profile_photo')
            if profile_photo and allowed_file(profile_photo.filename):
                # Delete old photo if exists
                if employee.profile_photo:
                    old_photo_path = os.path.join('static', employee.profile_photo)
                    if os.path.exists(old_photo_path):
                        os.remove(old_photo_path)
                
                # Save new photo
                filename = secure_filename(profile_photo.filename)
                photos_dir = os.path.join('static', 'employee_photos')
                if not os.path.exists(photos_dir):
                    os.makedirs(photos_dir)
                profile_photo_path = os.path.join('employee_photos', filename)
                profile_photo.save(os.path.join('static', profile_photo_path))
                employee.profile_photo = profile_photo_path
            
            db.session.commit()
            flash('Employee updated successfully')
            return redirect(url_for('employee_dashboard'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating employee: {str(e)}')
    
    return render_template('edit_employee.html', 
                         employee=employee,
                         user_name=session.get('user_name'))

@app.route('/delete-employee/<int:employee_id>', methods=['POST'])
def delete_employee(employee_id):
    try:
        employee = Employee.query.get_or_404(employee_id)
        
        # Delete related records first
        Task.query.filter_by(employee_id=employee_id).delete()
        Attendance.query.filter_by(employee_id=employee_id).delete()
        Document.query.filter_by(employee_id=employee_id).delete()
        SalaryComponent.query.filter_by(employee_id=employee_id).delete()
        EmployeeSalary.query.filter_by(employee_id=employee_id).delete()
        Lead.query.filter_by(assigned_to=employee_id).delete()
        Client.query.filter_by(assigned_to=employee_id).delete()
        Milestone.query.filter_by(employee_id=employee_id).delete()
        
        # Delete the employee
        db.session.delete(employee)
        db.session.commit()
        
        flash('Employee deleted successfully', 'success')
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Error deleting employee. Please try again.'}), 500

@app.route('/employee/add-task/<int:employee_id>', methods=['POST'])
def add_employee_task(employee_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    try:
        task = Task(
            title=request.form.get('title'),
            description=request.form.get('description'),
            due_date=datetime.strptime(request.form.get('due_date'), '%Y-%m-%d'),
            status=request.form.get('status', 'pending'),
            employee_id=employee_id,
            assigned_by_id=session['user_id']
        )
        db.session.add(task)
        db.session.commit()
        flash('Task added successfully')
    except Exception as e:
        flash(f'Error adding task: {str(e)}')
    
    return redirect(url_for('employee_profile', employee_id=employee_id, active_tab='tasks'))

@app.route('/employee/update-task/<int:task_id>', methods=['POST'])
def update_employee_task(task_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    task = Task.query.get_or_404(task_id)
    try:
        # Handle status-only updates
        if 'status' in request.form and len(request.form) == 1:
            task.status = request.form.get('status')
        else:
            # Handle full task updates
            task.title = request.form.get('title')
            task.description = request.form.get('description')
            task.due_date = datetime.strptime(request.form.get('due_date'), '%Y-%m-%d')
            task.status = request.form.get('status', task.status)
            task.priority = request.form.get('priority', task.priority)
            if request.form.get('employee_id'):
                task.employee_id = request.form.get('employee_id')
        
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/employee/get-task/<int:task_id>')
def get_employee_task(task_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
        
    task = Task.query.get_or_404(task_id)
    return jsonify({
        'title': task.title,
        'description': task.description,
        'due_date': task.due_date.strftime('%Y-%m-%d'),
        'status': task.status,
        'employee_id': task.employee_id,
        'priority': task.priority
    })

@app.route('/employee/delete-task/<int:task_id>', methods=['POST'])
def delete_employee_task(task_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
        
    task = Task.query.get_or_404(task_id)
    try:
        db.session.delete(task)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/add-milestone/<int:employee_id>', methods=['POST'])
def add_milestone(employee_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    try:
        milestone = Milestone(
            title=request.form.get('title'),
            description=request.form.get('description'),
            date=datetime.strptime(request.form.get('date'), '%Y-%m-%d'),
            employee_id=employee_id,
            created_by=session['user_id']
        )
        db.session.add(milestone)
        db.session.commit()
        flash('Milestone added successfully')
    except Exception as e:
        flash(f'Error adding milestone: {str(e)}')
    
    return redirect(url_for('employee_profile', employee_id=employee_id))

@app.route('/edit-milestone/<int:milestone_id>', methods=['POST'])
def edit_milestone(milestone_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    milestone = Milestone.query.get_or_404(milestone_id)
    try:
        milestone.title = request.form.get('title')
        milestone.description = request.form.get('description')
        milestone.date = datetime.strptime(request.form.get('date'), '%Y-%m-%d')
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/delete-milestone/<int:milestone_id>', methods=['POST'])
def delete_milestone(milestone_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    milestone = Milestone.query.get_or_404(milestone_id)
    try:
        db.session.delete(milestone)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get-milestone/<int:milestone_id>')
def get_milestone(milestone_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    milestone = Milestone.query.get_or_404(milestone_id)
    return jsonify({
        'title': milestone.title,
        'description': milestone.description,
        'date': milestone.date.strftime('%Y-%m-%d')
    })

@app.route('/api/mark-attendance/<int:employee_id>', methods=['POST'])
def add_attendance_old(employee_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    try:
        # Check if attendance already exists for this date
        attendance_date = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
        existing_attendance = Attendance.query.filter_by(
            employee_id=employee_id,
            date=attendance_date
        ).first()
        
        if existing_attendance:
            # Update existing attendance
            existing_attendance.status = request.form.get('status')
            existing_attendance.comments = request.form.get('comments', '')
            flash('Attendance updated successfully')
        else:
            # Create new attendance record
            attendance = Attendance(
                employee_id=employee_id,
                date=attendance_date,
                status=request.form.get('status'),
                comments=request.form.get('comments', '')
            )
            db.session.add(attendance)
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Attendance marked successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# Excel import/export functions
@app.route('/upload-employees', methods=['POST'])
def upload_employees():
    if 'file' not in request.files:
        flash('No file uploaded', 'error')
        return redirect(url_for('employee_dashboard'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('employee_dashboard'))
    
    if not file.filename.endswith('.xlsx'):
        flash('Please upload an Excel file (.xlsx)', 'error')
        return redirect(url_for('employee_dashboard'))
    
    try:
        df = pd.read_excel(file)
        required_columns = ['Name', 'Email', 'Phone', 'Role', 'Department', 'Employee Type', 'Salary', 'Join Date']
        
        # Check if all required columns are present
        if not all(col in df.columns for col in required_columns):
            flash('Invalid template format. Please use the provided template.', 'error')
            return redirect(url_for('employee_dashboard'))
        
        success_count = 0
        error_count = 0
        
        for _, row in df.iterrows():
            try:
                # Convert salary to float
                salary = float(row['Salary']) if pd.notna(row['Salary']) else 0.0
                
                # Convert join date to datetime
                join_date = pd.to_datetime(row['Join Date']).date() if pd.notna(row['Join Date']) else datetime.now().date()
                
                # Check if employee with email already exists
                existing_employee = Employee.query.filter_by(email=row['Email']).first()
                if existing_employee:
                    error_count += 1
                    continue
                
                new_employee = Employee(
                    name=row['Name'],
                    email=row['Email'],
                    phone=str(row['Phone']),
                    role=row['Role'],
                    department=row['Department'],
                    employee_type=row['Employee Type'],
                    salary=salary,
                    join_date=join_date
                )
                db.session.add(new_employee)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                continue
        
        db.session.commit()
        flash(f'Successfully added {success_count} employees. {error_count} entries had errors.', 'success')
        
    except Exception as e:
        flash('Error processing file. Please ensure you are using the correct template.', 'error')
        
    return redirect(url_for('employee_dashboard'))

@app.route('/export-employees')
def export_employees():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    employees = Employee.query.all()
    data = []
    for emp in employees:
        data.append({
            'name': emp.name,
            'email': emp.email,
            'phone': emp.phone,
            'dob': emp.dob.strftime('%Y-%m-%d') if emp.dob else '',
            'role': emp.role,
            'department': emp.department,
            'employee_type': emp.employee_type,
            'salary': emp.salary,
            'pay_frequency': emp.pay_frequency,
            'join_date': emp.join_date.strftime('%Y-%m-%d')
        })
    
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)
    
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='employees.xlsx'
    )

@app.route('/download-employee-template')
def download_employee_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Employee Template'
    
    # Define headers
    headers = ['Name', 'Email', 'Phone', 'Role', 'Department', 'Employee Type', 'Salary', 'Join Date']
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col, value=header)
    
    # Add sample data
    sample_data = [
        'John Doe', 'john@example.com', '1234567890', 'Developer', 'IT', 'Full-time', '50000', '2024-01-01'
    ]
    for col, value in enumerate(sample_data, 1):
        sheet.cell(row=2, column=col, value=value)
    
    # Save to BytesIO
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='employee_template.xlsx'
    )

@app.route('/export-candidates')
def export_candidates():
    candidates = Candidate.query.all()
    data = []
    for candidate in candidates:
        data.append({
            'name': candidate.name,
            'email': candidate.email,
            'phone': candidate.phone,
            'role_applied': candidate.role_applied,
            'experience': candidate.experience,
            'status': candidate.status,
            'created_date': candidate.created_date.strftime('%Y-%m-%d')
        })
    
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)
    
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='candidates.xlsx'
    )

@app.route('/download-candidate-template')
def download_candidate_template():
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"
    
    # Define headers - matching exactly with the upload processing
    headers = ['name', 'email', 'phone', 'role_applied', 'experience', 'source']
    ws.append(headers)
    
    # Add sample data
    sample_data = [
        'John Doe',
        'john@example.com',
        '1234567890',
        'Software Engineer',
        '5',
        'LinkedIn'
    ]
    ws.append(sample_data)
    
    # Create a temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False)
    wb.save(temp_file.name)
    
    return send_file(
        temp_file.name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='candidate_template.xlsx'
    )

@app.route('/upload-candidates', methods=['POST'])
def upload_candidates():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if 'file' not in request.files:
        flash('No file uploaded')
        return redirect(url_for('recruitment_dashboard'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected')
        return redirect(url_for('recruitment_dashboard'))
    
    if not file.filename.endswith('.xlsx'):
        flash('Invalid file format. Please upload an Excel file')
        return redirect(url_for('recruitment_dashboard'))
    
    try:
        df = pd.read_excel(file)
        required_columns = ['name', 'email', 'phone', 'role_applied', 'experience', 'source']
        
        if not all(col in df.columns for col in required_columns):
            flash('Invalid file format. Please use the template provided')
            return redirect(url_for('recruitment_dashboard'))
        
        success_count = 0
        error_count = 0
        
        for _, row in df.iterrows():
            try:
                # Convert experience to float and handle NaN
                experience = float(row['experience']) if pd.notna(row['experience']) else 0.0
                
                candidate = Candidate(
                    name=str(row['name']),
                    email=str(row['email']),
                    phone=str(row['phone']),
                    role_applied=str(row['role_applied']),
                    experience=experience,
                    source=str(row['source']),
                    status='Applied',
                    created_date=datetime.utcnow()
                )
                db.session.add(candidate)
                success_count += 1
            except Exception as e:
                error_count += 1
                continue
        
        db.session.commit()
        flash(f'Successfully added {success_count} candidates. {error_count} entries had errors.')
    except Exception as e:
        flash(f'Error uploading candidates: {str(e)}')
    
    return redirect(url_for('recruitment_dashboard'))

@app.route('/get-candidate/<int:candidate_id>')
def get_candidate(candidate_id):
    if 'user_id' not in session:
        return jsonify({'error': 'Not authorized'}), 401
    
    candidate = Candidate.query.get_or_404(candidate_id)
    return jsonify({
        'name': candidate.name,
        'email': candidate.email,
        'phone': candidate.phone,
        'position': candidate.role_applied,
        'experience': candidate.experience,
        'status': candidate.status,
        'has_resume': bool(candidate.resume_path)
    })

@app.route('/update-candidate/<int:candidate_id>', methods=['POST'])
def update_candidate(candidate_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    candidate = Candidate.query.get_or_404(candidate_id)
    try:
        candidate.name = request.form.get('name')
        candidate.email = request.form.get('email')
        candidate.phone = request.form.get('phone')
        candidate.role_applied = request.form.get('position')
        candidate.status = request.form.get('status')
        candidate.experience = float(request.form.get('experience', 0))
        
        # Handle resume update if provided
        resume = request.files.get('resume')
        if resume and allowed_file(resume.filename):
            # Delete old resume if exists
            if candidate.resume_path and os.path.exists(candidate.resume_path):
                os.remove(candidate.resume_path)
            
            filename = secure_filename(f"{int(time.time())}_{resume.filename}")
            resume_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            resume.save(resume_path)
            candidate.resume_path = resume_path
        
        db.session.commit()
        flash('Candidate updated successfully')
    except Exception as e:
        flash(f'Error updating candidate: {str(e)}')
    
    return redirect(url_for('recruitment_dashboard'))

@app.route('/delete-candidate/<int:candidate_id>', methods=['POST'])
def delete_candidate(candidate_id):
    candidate = Candidate.query.get_or_404(candidate_id)
    try:
        if candidate.resume_path and os.path.exists(candidate.resume_path):
            os.remove(candidate.resume_path)
        db.session.delete(candidate)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get-resume/<int:candidate_id>')
def get_resume(candidate_id):
    if 'user_id' not in session:
        return jsonify({'error': 'Not authorized'}), 401
            
    candidate = Candidate.query.get_or_404(candidate_id)
    
    try:
        if candidate.resume_path and os.path.exists(candidate.resume_path):
            # Get the file extension to set correct MIME type
            file_ext = os.path.splitext(candidate.resume_path)[1].lower()
            mime_type = {
                '.pdf': 'application/pdf',
                '.doc': 'application/msword',
                '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            }.get(file_ext, 'application/octet-stream')
            
            return send_file(
                candidate.resume_path,
                mimetype=mime_type,
                as_attachment=False,  # This will display in browser instead of downloading
                download_name=os.path.basename(candidate.resume_path)
            )
        else:
            return jsonify({'error': 'Resume not found'}), 404
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        try:
            hashed_password = generate_password_hash(request.form.get('password'))
            user = User(
                name=request.form.get('name'),
                email=request.form.get('email'),
                password=hashed_password
            )
            db.session.add(user)
            db.session.commit()
            flash('Account created successfully')
            return redirect(url_for('login'))
        except Exception as e:
            flash(f'Error creating account: {str(e)}')
    
    return render_template('signup.html')

@app.route('/document-finder')
def document_finder():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get current user and check admin status
    current_user = User.query.get(session['user_id'])
    is_admin = current_user.email == 'admin@example.com'
    has_confidential_access = is_admin or session.get('confidential_access', False)
    
    # Get show_confidential parameter
    show_confidential = request.args.get('confidential') == 'true' and has_confidential_access
    
    # Get all document categories
    categories = DocumentCategory.query.all()
    
    # Filter documents based on search, category, and confidential access
    query = Document.query
    
    # Apply search filter if provided
    search_term = request.args.get('search')
    if search_term:
        query = query.filter(
            db.or_(
                Document.name.ilike(f'%{search_term}%'),
                Document.description.ilike(f'%{search_term}%')
            )
        )
    
    # Apply category filter if provided
    category_id = request.args.get('category')
    if category_id and category_id != 'all':
        query = query.filter(Document.category_id == category_id)
    
    # Filter based on confidential access
    if show_confidential:
        query = query.join(DocumentCategory).filter(DocumentCategory.is_confidential == True)
    else:
        query = query.join(DocumentCategory).filter(DocumentCategory.is_confidential == False)
    
    # Get documents with category information
    documents = query.order_by(Document.uploaded_date.desc()).all()
    
    return render_template('document_finder.html',
                         documents=documents,
                         categories=categories,
                         user_name=session.get('user_name'),
                         is_admin=is_admin,
                         has_confidential_access=has_confidential_access,
                         show_confidential=show_confidential)

@app.route('/upload_document', methods=['POST'])
@app.route('/upload_document/<int:employee_id>', methods=['POST'])
def upload_document(employee_id=None):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    try:
        # Get current user
        current_user = User.query.get(session.get('user_id'))
        if not current_user:
            flash('Please log in again', 'error')
            return redirect(url_for('login'))

        # Validate file upload
        if 'document' not in request.files:
            flash('No document provided', 'error')
            return redirect(url_for('employee_profile', employee_id=employee_id) if employee_id else url_for('document_finder'))
        
        file = request.files['document']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('employee_profile', employee_id=employee_id) if employee_id else url_for('document_finder'))
        
        if not allowed_file(file.filename):
            flash('File type not allowed', 'error')
            return redirect(url_for('employee_profile', employee_id=employee_id) if employee_id else url_for('document_finder'))
        
        # Process file upload
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{timestamp}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        
        # Save the file
        file.save(file_path)
        
        # Create document record
        new_document = Document(
            name=request.form.get('name', filename),
            file_path=file_path,
            document_type=request.form.get('document_type', 'personal'),
            category_id=request.form.get('category_id'),
            description=request.form.get('description'),
            uploaded_by=current_user.id,
            employee_id=employee_id,
            uploaded_date=datetime.utcnow()
        )
        
        db.session.add(new_document)
        db.session.commit()
        
        flash('Document uploaded successfully', 'success')
        
        # Always redirect to employee profile if employee_id is provided
        if employee_id:
            return redirect(url_for('employee_profile', employee_id=employee_id))
        return redirect(url_for('document_finder'))
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error uploading document: {str(e)}', 'error')
        if employee_id:
            return redirect(url_for('employee_profile', employee_id=employee_id))
        return redirect(url_for('document_finder'))

def calculate_change(current, previous):
    if previous == 0:
        return 100 if current > 0 else 0
    return ((current - previous) / previous) * 100

@app.route('/verify-confidential-access', methods=['POST'])
def verify_confidential_access():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    data = request.get_json()
    access_code = data.get('access_code')
    if not access_code:
        return jsonify({'success': False, 'error': 'Access code is required'}), 400
    current_user = User.query.get(session['user_id'])
    if current_user.email == 'admin@example.com' or access_code == 'EQ#2024@SEC$KEY':
        session['confidential_access'] = True
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'Invalid access code'}), 401

@app.route('/feedback-management')
def feedback_management():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get current user
    current_user = User.query.get(session['user_id'])
    
    # Get all employees for the dropdown
    employees = Employee.query.all()
    
    # Get feedback given by current user
    given_feedback = EmployeeFeedback.query.filter_by(reviewer_id=session['user_id']).all()
    
    # Get submitted client feedback (where rating is not null)
    client_feedback = ClientFeedback.query.filter(
        ClientFeedback.rating.isnot(None)
    ).order_by(ClientFeedback.created_date.desc()).all()
    
    # Get pending feedback links (where rating is null)
    pending_feedback_links = ClientFeedback.query.filter(
        ClientFeedback.rating.is_(None)
    ).order_by(ClientFeedback.feedback_link_id.desc()).all()
    
    return render_template('feedback_management.html',
                         employees=employees,
                         given_feedback=given_feedback,
                         client_feedback=client_feedback,
                         pending_feedback_links=pending_feedback_links,
                         user_name=session.get('user_name'))

@app.route('/leave-management')
def leave_management():
    if 'user_id' not in session:
        return redirect(url_for('login'))
        
    # Get all leave types
    leave_types = LeaveType.query.all()
    
    # Get leave requests made by the current user
    my_requests = LeaveRequest.query.filter_by(employee_id=session['user_id']).order_by(LeaveRequest.created_at.desc()).all()
    
    # Get pending approvals only if user is admin (user_id = 1)
    pending_approvals = []
    if session['user_id'] == 1:
        pending_approvals = LeaveRequest.query.filter_by(
            approver_id=1,
            status='pending'
        ).order_by(LeaveRequest.created_at.desc()).all()
    
    return render_template('leave_management.html',
                         leave_types=leave_types,
                         my_requests=my_requests,
                         pending_approvals=pending_approvals,
                         current_user_id=session['user_id'])

@app.route('/submit-leave-request', methods=['POST'])
def submit_leave_request():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    try:
        leave_request = LeaveRequest(
            employee_id=session['user_id'],
            leave_type_id=request.form.get('leave_type'),
            start_date=datetime.strptime(request.form.get('start_date'), '%Y-%m-%d'),
            end_date=datetime.strptime(request.form.get('end_date'), '%Y-%m-%d'),
            reason=request.form.get('reason'),
            status='pending',
            approver_id=request.form.get('approver_id')
        )
        
        db.session.add(leave_request)
        db.session.commit()
        
        flash('Leave request submitted successfully')
    except Exception as e:
        flash(f'Error submitting leave request: {str(e)}')
    
    return redirect(url_for('leave_management'))

@app.route('/lead-client-management')
def lead_client_management():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    leads = Lead.query.all()
    clients = Client.query.all()
    employees = Employee.query.all()  # Get all employees for assignment
    
    return render_template('lead_client_management.html',
                         leads=leads,
                         clients=clients,
                         employees=employees,
                         user_name=session.get('user_name'))


@app.route('/delete-salary-component/<int:component_id>', methods=['POST'])
def delete_salary_component(component_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Not logged in'}), 401
    
    try:
        # Get the component
        component = SalaryComponent.query.get_or_404(component_id)
        
        # Delete any salary details that reference this component
        SalaryDetail.query.filter_by(component_id=component_id).delete()
        
        # Delete the component
        db.session.delete(component)
        db.session.commit()
        
        flash('Salary component deleted successfully', 'success')
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting salary component: {str(e)}', 'error')
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/salary-management')
def salary_management():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    selected_month = int(request.args.get('month', datetime.now().month))
    selected_year = int(request.args.get('year', datetime.now().year))

    employees = Employee.query.all()
    employee_salaries = []
    total_payroll = 0
    pending_count = 0
    processed_count = 0

    for employee in employees:
        # Get employee's salary record for selected month
        salary = EmployeeSalary.query.filter_by(
            employee_id=employee.id,
            month=selected_month,
            year=selected_year
        ).first()

        # Get salary components
        components = SalaryComponent.query.filter_by(employee_id=employee.id).all()
        
        # Calculate net salary based on basic pay and components
        basic_pay = employee.salary or 0  # Get basic pay from employee profile
        net_salary = basic_pay
        
        # Add component amounts to net salary
        for component in components:
            amount = component.value
            if component.is_percentage:
                amount = (basic_pay * component.value) / 100
            
            if component.type == 'allowance':
                net_salary += amount
            elif component.type == 'deduction':
                net_salary -= amount

        # If salary record exists, use its values
        if salary:
            basic_pay = salary.basic_pay
            net_salary = salary.net_salary or net_salary
            status = salary.status
            salary_id = salary.id
        else:
            status = 'not_generated'
            salary_id = None

        # Update counters
        if status == 'pending':
            pending_count += 1
        elif status == 'processed':
            processed_count += 1
            total_payroll += net_salary

        # Add to employee salaries list
        employee_salaries.append({
            'employee': employee,
            'id': salary_id,
            'basic_pay': basic_pay,
            'net_salary': net_salary,
            'status': status,
            'components': components
        })

    return render_template('salary_management.html',
                         employee_salaries=employee_salaries,
                         total_payroll=total_payroll,
                         pending_count=pending_count,
                         processed_count=processed_count,
                         selected_month=selected_month,
                         selected_year=selected_year,
                         employees=employees,
                         user_name=session.get('user_name'))

@app.route('/task-productivity')
def task_productivity():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get current user
    current_user = User.query.get(session['user_id'])
    is_admin = current_user.email == 'admin@example.com'
    
    # Get all employees
    employees = Employee.query.all()

    # Get tasks for each employee
    employee_tasks = {}
    employee_productivity = {}
    
    # Initialize completion dates list
    today = datetime.now()
    thirty_days_ago = today - timedelta(days=30)
    
    completion_dates = []
    completion_counts = []
    
    # Get data for last 30 days
    current_date = thirty_days_ago
    while current_date <= today:
        date_str = current_date.strftime('%Y-%m-%d')
        completion_dates.append(date_str)
        
        # Count tasks completed on this date
        completed_count = Task.query.filter(
            Task.status == 'completed',
            func.date(Task.created_date) == current_date.date()
        ).count()
        completion_counts.append(completed_count)
        current_date += timedelta(days=1)

    # Prepare overdue tasks data
    overdue_tasks_data = []

    # Get employee data for tasks
    for employee in employees:
        tasks = Task.query.filter_by(employee_id=employee.id).all()
        
        serialized_tasks = []
        for task in tasks:
            serialized_task = {
                'id': task.id,
                'title': task.title,
                'description': task.description,
                'due_date': task.due_date.strftime('%Y-%m-%d') if task.due_date else None,
                'priority': task.priority,
                'status': task.status,
                'created_date': task.created_date.strftime('%Y-%m-%d') if task.created_date else None
            }
            serialized_tasks.append(serialized_task)
        
        employee_tasks[str(employee.id)] = serialized_tasks
        
        # Calculate productivity score
        total_tasks = len(tasks)
        if total_tasks > 0:
            completed_tasks = len([t for t in tasks if t.status == 'completed'])
            productivity = round((completed_tasks / total_tasks) * 100, 2)
        else:
            productivity = 0
        
        employee_productivity[str(employee.id)] = productivity

        # Count overdue tasks
        overdue_count = Task.query.filter(
            Task.employee_id == employee.id,
            Task.status != 'completed',
            Task.due_date < datetime.now()
        ).count()
    
        if overdue_count > 0:
            overdue_tasks_data.append({
                'name': employee.name,
                'count': overdue_count
            })
    
    return render_template('task_productivity.html',
                         employees=employees,
                         employee_tasks=employee_tasks,
                         employee_productivity=employee_productivity,
                         completion_dates=completion_dates,
                         completion_counts=completion_counts,
                         overdue_tasks_data=overdue_tasks_data,
                         user_name=session.get('user_name'),
                         is_admin=is_admin)

@app.route('/analytics-dashboard')
def analytics_dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get current user
    current_user = User.query.get(session['user_id'])
    
    # Get all employees
    employees = Employee.query.all()
    
    # Calculate employee metrics
    total_employees = len(employees)
    active_employees = len([e for e in employees if e.employee_type != 'terminated'])
    
    # Task metrics
    tasks = Task.query.all()
    completed_tasks = len([t for t in tasks if t.status == 'completed'])
    pending_tasks = len([t for t in tasks if t.status == 'pending'])
    
    # Attendance metrics
    today = datetime.now()
    month_start = date(today.year, today.month, 1)
    attendance_records = Attendance.query.filter(
        Attendance.date >= month_start
    ).all()

    present_count = len([a for a in attendance_records if a.status == 'present'])
    absent_count = len([a for a in attendance_records if a.status == 'absent'])
    
    return render_template('analytics_dashboard.html',
                         total_employees=total_employees,
                         active_employees=active_employees,
                         completed_tasks=completed_tasks,
                         pending_tasks=pending_tasks,
                         present_count=present_count,
                         absent_count=absent_count,
                         user_name=session.get('user_name'))
            
def get_task_analytics():
    try:
        tasks = Task.query.all()
        total = len(tasks)
        completed_tasks = sum(1 for t in tasks if t.status == 'Completed')
        pending = sum(1 for t in tasks if t.status == 'In Progress')
        overdue = sum(1 for t in tasks if t.due_date and t.due_date < datetime.now() and t.status != 'Completed')
        return total, completed_tasks, pending, overdue
    except Exception as e:
        print(f"Error in task analytics: {str(e)}")
        return 0, 0, 0, 0

def get_employee_analytics():
    try:
        employees = Employee.query.all()
        total = len(employees)
        roles = len(set(emp.role for emp in employees if emp.role))
        return total, roles
    except Exception as e:
        print(f"Error in employee analytics: {str(e)}")
        return 0, 0

def get_recruitment_analytics():
    try:
        candidates = Candidate.query.all()
        candidates_applied = len(candidates)  # Total number of candidates who applied
        selected = sum(1 for c in candidates if c.status == 'Selected')
        return candidates_applied, selected
    except Exception as e:
        print(f"Error in recruitment analytics: {str(e)}")
        return 0, 0

def get_budget_analytics():
    try:
        departments = Department.query.all()
        total = sum(d.budget for d in departments)
        expenses = Expense.query.all()
        spent = sum(e.amount for e in expenses)
        recent = sum(e.amount for e in expenses if e.date >= (datetime.now() - timedelta(days=30)))
        return total, total - spent, recent
    except Exception as e:
        print(f"Error in budget analytics: {str(e)}")
        return 0, 0, 0

def get_client_analytics():
    try:
        clients = Client.query.all()
        leads = Lead.query.all()
        active = sum(1 for c in clients if c.status == 'active')
        new = sum(1 for l in leads if l.status == 'new')
        return active, new
    except Exception as e:
        print(f"Error in client analytics: {str(e)}")
        return 0, 0

def get_social_analytics():
    try:
        campaigns = SocialMediaCampaign.query.all()
        posts = ScheduledPost.query.all()
        total_campaigns = len(campaigns)  # Total number of campaigns
        scheduled = sum(1 for p in posts if p.scheduled_time > datetime.now())
        spend = sum(c.amount_spent for c in campaigns if hasattr(c, 'amount_spent') and c.amount_spent is not None)
        return total_campaigns, scheduled, spend
    except Exception as e:
        print(f"Error in social analytics: {str(e)}")
        return 0, 0, 0

@app.route('/analytics')
def analytics():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    # Get analytics from each section
    total_tasks, completed_tasks, pending_tasks, overdue_tasks = get_task_analytics()
    total_employees, total_roles = get_employee_analytics()
    candidates_applied, selected_candidates = get_recruitment_analytics()
    total_budget, remaining_budget, recent_expenses = get_budget_analytics()
    active_clients, new_leads = get_client_analytics()
    total_campaigns, scheduled_posts, campaign_spend = get_social_analytics()

    return render_template('analytics_dashboard.html',
                        total_tasks=total_tasks,
                        completed_tasks=completed_tasks,
                        pending_tasks=pending_tasks,
                        overdue_tasks=overdue_tasks,
                        total_employees=total_employees,
                        total_roles=total_roles,
                        candidates_applied=candidates_applied,
                        selected_candidates=selected_candidates,
                        total_budget=total_budget,
                        remaining_budget=remaining_budget,
                        recent_expenses=recent_expenses,
                        active_clients=active_clients,
                        new_leads=new_leads,
                        total_campaigns=total_campaigns,
                        scheduled_posts=scheduled_posts,
                        campaign_spend=campaign_spend,
                        user_name=session.get('user_name'))

@app.route('/submit-employee-feedback', methods=['POST'])
def submit_employee_feedback():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    try:
        feedback = EmployeeFeedback(
            employee_id=request.form.get('employee_id'),
            reviewer_id=session['user_id'],
            feedback_type=request.form.get('feedback_type'),
            rating=int(request.form.get('rating')),
            comments=request.form.get('comments')
        )
        db.session.add(feedback)
        db.session.commit()
        flash('Feedback submitted successfully')
    except Exception as e:
        flash(f'Error submitting feedback: {str(e)}')
    
    return redirect(url_for('feedback_management'))

@app.route('/submit-client-feedback/<string:link_id>', methods=['POST'])
def submit_client_feedback(link_id):
    feedback = ClientFeedback.query.filter_by(feedback_link_id=link_id).first_or_404()
    
    if feedback.rating is not None:
        flash('This feedback link has already been used')
        return redirect(url_for('feedback_submitted'))
    
    try:
        feedback.rating = int(request.form.get('rating'))
        feedback.comments = request.form.get('comments')
        feedback.created_date = datetime.utcnow()
        db.session.commit()
        flash('Thank you for your feedback!')
    except Exception as e:
        flash(f'Error submitting feedback: {str(e)}')
    
    return redirect(url_for('feedback_submitted'))

@app.route('/manage-document-categories')
def manage_document_categories():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Check if user is admin
    current_user = User.query.get(session['user_id'])
    if current_user.email != 'admin@example.com':
        flash('Access denied. Only admin can manage document categories.')
        return redirect(url_for('document_finder'))
    
    categories = DocumentCategory.query.all()
    return render_template('manage_document_categories.html',
                         categories=categories,
                         user_name=session.get('user_name'))

@app.route('/add-document-category', methods=['POST'])
def add_document_category():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    # Check if user is admin
    current_user = User.query.get(session['user_id'])
    if current_user.email != 'admin@example.com':
        flash('Access denied. Only admin can add document categories.')
        return redirect(url_for('document_finder'))
    
    try:
        category = DocumentCategory(
            name=request.form.get('name'),
            is_confidential=request.form.get('is_confidential') == 'true',
            access_code=request.form.get('access_code') if request.form.get('is_confidential') == 'true' else None
        )
        db.session.add(category)
        db.session.commit()
        flash('Category added successfully')
    except Exception as e:
        flash(f'Error adding category: {str(e)}')
    
    return redirect(url_for('manage_document_categories'))

@app.route('/feedback-submitted')
def feedback_submitted():
    return render_template('feedback_submitted.html')

@app.route('/download-document/<int:document_id>')
def download_document(document_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    document = Document.query.get_or_404(document_id)
    try:
        # Get file extension
        file_ext = os.path.splitext(document.file_path)[1].lower()
        
        return send_file(
            document.file_path,
            as_attachment=True,
            download_name=f"{document.name}{file_ext}",
            mimetype=get_mimetype(file_ext)
        )
    except Exception as e:
        flash(f'Error downloading document: {str(e)}', 'error')
        return redirect(url_for('employee_profile', employee_id=document.employee_id))

@app.route('/view-document/<int:document_id>')
def view_document(document_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    document = Document.query.get_or_404(document_id)
    try:
        # Get file extension
        file_ext = os.path.splitext(document.file_path)[1].lower()
        
        return send_file(
            document.file_path,
            as_attachment=False,
            mimetype=get_mimetype(file_ext)
        )
    except Exception as e:
        flash(f'Error viewing document: {str(e)}', 'error')
        return redirect(url_for('employee_profile', employee_id=document.employee_id))

def get_mimetype(file_ext):
    """Helper function to get correct mimetype for file extensions"""
    mimetypes = {
        '.pdf': 'application/pdf',
        '.doc': 'application/msword',
        '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        '.xls': 'application/vnd.ms-excel',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png',
        '.txt': 'text/plain'
    }
    return mimetypes.get(file_ext, 'application/octet-stream')

@app.route('/delete-document/<int:document_id>', methods=['POST'])
def delete_document(document_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    document = Document.query.get_or_404(document_id)
    try:
        # Delete the physical file
        if os.path.exists(document.file_path):
            os.remove(document.file_path)
        
        # Delete the database record
        db.session.delete(document)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get-campaign/<int:campaign_id>')
def get_campaign(campaign_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    campaign = SocialMediaCampaign.query.get_or_404(campaign_id)
    return jsonify({
        'name': campaign.name,
        'budget': campaign.budget,
        'amount_spent': campaign.amount_spent,
        'start_date': campaign.start_date.strftime('%Y-%m-%d'),
        'end_date': campaign.end_date.strftime('%Y-%m-%d'),
        'status': campaign.status
    })

@app.route('/update-campaign/<int:campaign_id>', methods=['POST'])
def update_campaign(campaign_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    campaign = SocialMediaCampaign.query.get_or_404(campaign_id)
    try:
        campaign.name = request.form.get('name')
        campaign.budget = float(request.form.get('budget'))
        campaign.amount_spent = float(request.form.get('amount_spent'))
        campaign.start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d')
        campaign.end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d')
        campaign.status = request.form.get('status')
        
        db.session.commit()
        flash('Campaign updated successfully')
    except Exception as e:
        flash(f'Error updating campaign: {str(e)}')
    
    return redirect(url_for('social_media_dashboard'))

@app.route('/delete-campaign/<int:campaign_id>', methods=['POST'])
def delete_campaign(campaign_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    campaign = SocialMediaCampaign.query.get_or_404(campaign_id)
    try:
        # Delete associated posts first
        ScheduledPost.query.filter_by(campaign_id=campaign_id).delete()
        # Delete the campaign
        db.session.delete(campaign)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get-post/<int:post_id>')
def get_post(post_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    post = ScheduledPost.query.get_or_404(post_id)
    return jsonify({
        'content': post.content,
        'platforms': post.platforms,
        'scheduled_time': post.scheduled_time.strftime('%Y-%m-%dT%H:%M'),
        'campaign_id': post.campaign_id
    })

@app.route('/update-post/<int:post_id>', methods=['POST'])
def update_post(post_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    post = ScheduledPost.query.get_or_404(post_id)
    try:
        platforms = ','.join(request.form.getlist('platforms'))
        post.content = request.form.get('content')
        post.platforms = platforms
        post.scheduled_time = datetime.strptime(request.form.get('scheduled_time'), '%Y-%m-%dT%H:%M')
        
        db.session.commit()
        flash('Post updated successfully')
    except Exception as e:
        flash(f'Error updating post: {str(e)}')
    
    return redirect(url_for('social_media_dashboard'))

@app.route('/delete-post/<int:post_id>', methods=['POST'])
def delete_post(post_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    post = ScheduledPost.query.get_or_404(post_id)
    try:
        db.session.delete(post)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get-department/<int:department_id>')
def get_department(department_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    department = Department.query.get_or_404(department_id)
    return jsonify({
        'name': department.name,
        'budget': department.budget
    })

@app.route('/update-department/<int:department_id>', methods=['POST'])
def update_department(department_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    department = Department.query.get_or_404(department_id)
    try:
        department.name = request.form.get('name')
        department.budget = float(request.form.get('budget'))
        db.session.commit()
        flash('Department updated successfully')
    except Exception as e:
        flash(f'Error updating department: {str(e)}')
    
    return redirect(url_for('budget_dashboard'))

@app.route('/delete-department/<int:department_id>', methods=['POST'])
def delete_department(department_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    department = Department.query.get_or_404(department_id)
    try:
        # Delete associated expenses first
        Expense.query.filter_by(department_id=department_id).delete()
        # Delete the department
        db.session.delete(department)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get-expense/<int:expense_id>')
def get_expense(expense_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    expense = Expense.query.get_or_404(expense_id)
    return jsonify({
        'description': expense.description,
        'amount': expense.amount,
        'department_id': expense.department_id,
        'date': expense.date.strftime('%Y-%m-%d')
    })

@app.route('/update-expense/<int:expense_id>', methods=['POST'])
def update_expense(expense_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    expense = Expense.query.get_or_404(expense_id)
    try:
        expense.description = request.form.get('description')
        expense.amount = float(request.form.get('amount'))
        expense.department_id = request.form.get('department_id')
        expense.date = datetime.strptime(request.form.get('date'), '%Y-%m-%d')
        
        # Handle receipt update if provided
        receipt = request.files.get('receipt')
        if receipt and allowed_file(receipt.filename):
            # Delete old receipt if exists
            if expense.receipt_path and os.path.exists(expense.receipt_path):
                os.remove(expense.receipt_path)
            
            filename = secure_filename(f"{int(time.time())}_{receipt.filename}")
            receipt_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            receipt.save(receipt_path)
            expense.receipt_path = receipt_path
        
        db.session.commit()
        flash('Expense updated successfully')
    except Exception as e:
        flash(f'Error updating expense: {str(e)}')
    
    return redirect(url_for('budget_dashboard'))

@app.route('/delete-expense/<int:expense_id>', methods=['POST'])
def delete_expense(expense_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    expense = Expense.query.get_or_404(expense_id)
    try:
        # Delete receipt file if exists
        if expense.receipt_path and os.path.exists(expense.receipt_path):
            os.remove(expense.receipt_path)
        db.session.delete(expense)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/view-receipt/<int:expense_id>')
def view_receipt(expense_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    expense = Expense.query.get_or_404(expense_id)
    
    if not expense.receipt_path:
        flash('No receipt found for this expense')
        return redirect(url_for('budget_dashboard'))
    
    try:
        full_path = os.path.join(app.config['UPLOAD_FOLDER'], expense.receipt_path)
        if not os.path.exists(full_path):
            flash('Receipt file not found')
            return redirect(url_for('budget_dashboard'))
        
        # Get file extension to determine content type
        _, ext = os.path.splitext(expense.receipt_path)
        if ext.lower() in ['.jpg', '.jpeg', '.png']:
            content_type = f'image/{ext[1:].lower()}'
        elif ext.lower() == '.pdf':
            content_type = 'application/pdf'
        else:
            content_type = 'application/octet-stream'
        
        return send_file(full_path, mimetype=content_type)
    except Exception as e:
        flash(f'Error viewing receipt: {str(e)}')
        return redirect(url_for('budget_dashboard'))

@app.route('/delete-document-category/<int:category_id>', methods=['POST'])
def delete_document_category(category_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    # Check if user is admin
    current_user = User.query.get(session['user_id'])
    if current_user.email != 'admin@example.com':
        return jsonify({'success': False, 'error': 'Access denied'})
    
    try:
        category = DocumentCategory.query.get_or_404(category_id)
        
        # Check if there are any documents in this category
        documents = Document.query.filter_by(category_id=category_id).all()
        
        # Delete all documents in this category first
        for document in documents:
            # Delete physical file if it exists
            if document.file_path and os.path.exists(document.file_path):
                os.remove(document.file_path)
            db.session.delete(document)
        
        # Now delete the category
        db.session.delete(category)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get-employee-feedback/<int:employee_id>')
def get_employee_feedback(employee_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    try:
        # Get all feedback received by this employee
        received_feedback = EmployeeFeedback.query.filter_by(employee_id=employee_id).all()
        
        feedback_list = []
        for feedback in received_feedback:
            reviewer = User.query.get(feedback.reviewer_id)
            feedback_list.append({
                'reviewer_name': reviewer.name,
                'feedback_type': feedback.feedback_type,
                'rating': feedback.rating,
                'comments': feedback.comments,
                'created_date': feedback.created_date.strftime('%Y-%m-%d')
            })
        
        return jsonify(feedback_list)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/generate-client-feedback-link', methods=['POST'])
def generate_client_feedback_link():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    try:
        # Generate a unique link ID
        link_id = ''.join(random.choices(string.ascii_letters + string.digits, k=16))
        
        # Create a new client feedback entry
        feedback = ClientFeedback(
            client_name=request.form.get('client_name'),
            client_email=request.form.get('client_email'),
            feedback_link_id=link_id,
            feedback_type='client'
        )
        
        db.session.add(feedback)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'link_id': link_id
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/client-feedback/<string:link_id>')
def client_feedback_form(link_id):
    feedback = ClientFeedback.query.filter_by(feedback_link_id=link_id).first_or_404()
    
    # Check if feedback has already been submitted
    if feedback.rating is not None:
        return render_template('feedback_submitted.html')
    
    return render_template('client_feedback_form.html', link_id=link_id)

@app.route('/apply-leave', methods=['POST'])
def apply_leave():
    if 'user_id' not in session:
        flash('Please login to apply for leave', 'error')
        return redirect(url_for('login'))

    try:
        # Get form data
        leave_type_id = request.form.get('leave_type')
        start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d')
        end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d')
        reason = request.form.get('reason')
        
        # Validate dates
        if start_date > end_date:
            flash('Start date cannot be after end date', 'error')
            return redirect(url_for('leave_management'))
            
        if start_date.date() < datetime.now().date():
            flash('Cannot apply leave for past dates', 'error')
            return redirect(url_for('leave_management'))
        
        # Create new leave request with admin as approver
        leave_request = LeaveRequest(
            employee_id=session['user_id'],
            leave_type_id=leave_type_id,
            start_date=start_date,
            end_date=end_date,
            reason=reason,
            status='pending',
            approver_id=1,  # Always set admin as approver
            created_at=datetime.now()
        )
        
        db.session.add(leave_request)
        db.session.commit()
        
        flash('Leave request submitted successfully! Pending admin approval.', 'success')
        return redirect(url_for('leave_management'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error submitting leave request: {str(e)}', 'error')
        return redirect(url_for('leave_management'))

@app.route('/approve-leave/<int:request_id>', methods=['POST'])
def approve_leave(request_id):
    if 'user_id' not in session or session['user_id'] != 1:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        leave_request = LeaveRequest.query.get_or_404(request_id)
        action = request.form.get('action')
        
        if action not in ['approve', 'reject']:
            return jsonify({'success': False, 'error': 'Invalid action'}), 400
        
        leave_request.status = 'approved' if action == 'approve' else 'rejected'
        leave_request.updated_at = datetime.utcnow()
        
        # Add a comment if provided
        comment = request.form.get('comment')
        if comment:
            leave_comment = LeaveComment(
                leave_request_id=request_id,
                user_id=session['user_id'],
                comment=comment
            )
            db.session.add(leave_comment)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Leave request {action}d successfully'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400

@app.route('/get-leave-calendar')
def get_leave_calendar():
    if 'user_id' not in session:
        return jsonify([])
    
    leave_requests = LeaveRequest.query.all()
    events = []
    
    for leave in leave_requests:
        color = {
            'pending': '#ffc107',
            'approved': '#28a745',
            'rejected': '#dc3545'
        }.get(leave.status, '#6c757d')
        
        events.append({
            'title': f"{leave.employee.name} - {leave.leave_type.name}",
            'start': leave.start_date.strftime('%Y-%m-%d'),
            'end': (leave.end_date + timedelta(days=1)).strftime('%Y-%m-%d'),
            'backgroundColor': color,
            'borderColor': color
        })
    
    return jsonify(events)

@app.route('/assign-lead/<int:lead_id>', methods=['POST'])
def assign_lead(lead_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    try:
        lead = Lead.query.get_or_404(lead_id)
        employee_id = request.form.get('assigned_to')
        
        if not employee_id:
            return jsonify({'success': False, 'error': 'No employee selected'})
            
        lead.assigned_to = employee_id
        lead.updated_date = datetime.utcnow()
        db.session.commit()
        
        employee = Employee.query.get(employee_id)
        return jsonify({
            'success': True,
            'message': f'Lead assigned to {employee.name}',
            'employee_name': employee.name
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/assign-client/<int:client_id>', methods=['POST'])
def assign_client(client_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    try:
        client = Client.query.get_or_404(client_id)
        employee_id = request.form.get('assigned_to')
        
        if not employee_id:
            return jsonify({'success': False, 'error': 'No employee selected'})
            
        client.assigned_to = employee_id
        client.updated_date = datetime.utcnow()
        db.session.commit()
        
        employee = Employee.query.get(employee_id)
        return jsonify({
            'success': True,
            'message': f'Client assigned to {employee.name}',
            'employee_name': employee.name
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/add-lead', methods=['POST'])
def add_lead():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    try:
        lead = Lead(
            name=request.form.get('name'),
            email=request.form.get('email'),
            phone=request.form.get('phone'),
            source=request.form.get('source'),
            status='new',
            score=0,
            assigned_to=request.form.get('assigned_to'),
            notes=request.form.get('notes')
        )
        db.session.add(lead)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Lead added successfully'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/add-client', methods=['POST'])
def add_client():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    try:
        client = Client(
            name=request.form.get('name'),
            email=request.form.get('email'),
            phone=request.form.get('phone'),
            company=request.form.get('company'),
            address=request.form.get('address'),
            status='active',
            assigned_to=request.form.get('assigned_to')
        )
        db.session.add(client)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Client added successfully'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get-lead/<int:lead_id>')
def get_lead(lead_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    try:
        lead = Lead.query.get_or_404(lead_id)
        return jsonify({
            'success': True,
            'name': lead.name,
            'email': lead.email,
            'phone': lead.phone,
            'source': lead.source,
            'status': lead.status,
            'score': lead.score,
            'assigned_to': lead.assigned_to,
            'notes': lead.notes
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get-client/<int:client_id>')
def get_client(client_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    try:
        client = Client.query.get_or_404(client_id)
        return jsonify({
            'success': True,
            'name': client.name,
            'email': client.email,
            'phone': client.phone,
            'company': client.company,
            'address': client.address,
            'status': client.status,
            'assigned_to': client.assigned_to
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/update-lead/<int:lead_id>', methods=['POST'])
def update_lead(lead_id):
    if not session.get('user_id'):
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        lead = Lead.query.get_or_404(lead_id)
        lead.name = request.form.get('name')
        lead.email = request.form.get('email')
        lead.phone = request.form.get('phone')
        lead.source = request.form.get('source')
        lead.status = request.form.get('status')
        lead.notes = request.form.get('notes')
        assigned_to = request.form.get('assigned_to')
        if assigned_to:
            lead.assigned_to = int(assigned_to)
        
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/delete-lead/<int:lead_id>', methods=['POST'])
def delete_lead(lead_id):
    if not session.get('user_id'):
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        lead = Lead.query.get_or_404(lead_id)
        db.session.delete(lead)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/upload_leads', methods=['POST'])
def upload_leads():
    if 'file' not in request.files:
        flash('No file selected', 'error')
        return redirect(url_for('lead_client_management'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('lead_client_management'))
    
    if not file.filename.endswith('.xlsx'):
        flash('Please upload an Excel file (.xlsx)', 'error')
        return redirect(url_for('lead_client_management'))
    
    try:
        df = pd.read_excel(file)
        required_columns = ['Name', 'Email', 'Phone', 'Source', 'Status', 'Notes']
        
        if not all(col in df.columns for col in required_columns):
            flash('Invalid template format. Please use the provided template.', 'error')
            return redirect(url_for('lead_client_management'))
        
        success_count = 0
        error_count = 0
        
        for _, row in df.iterrows():
            try:
                existing_lead = Lead.query.filter_by(email=row['Email']).first()
                if existing_lead:
                    error_count += 1
                    continue
                
                new_lead = Lead(
                    name=row['Name'],
                    email=row['Email'],
                    phone=str(row['Phone']),
                    source=row['Source'],
                    status=row['Status'],
                    notes=row['Notes'] if pd.notna(row['Notes']) else None
                )
                db.session.add(new_lead)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                continue
        
        db.session.commit()
        flash(f'Successfully added {success_count} leads. {error_count} leads failed.', 'success' if error_count == 0 else 'warning')
        
    except Exception as e:
        flash('Error processing file. Please check the format.', 'error')
    
    return redirect(url_for('lead_client_management'))

@app.route('/upload_clients', methods=['POST'])
def upload_clients():
    if 'file' not in request.files:
        flash('No file selected', 'error')
        return redirect(url_for('lead_client_management'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('lead_client_management'))
    
    if not file.filename.endswith('.xlsx'):
        flash('Please upload an Excel file (.xlsx)', 'error')
        return redirect(url_for('lead_client_management'))
    
    try:
        df = pd.read_excel(file)
        required_columns = ['Name', 'Email', 'Phone', 'Company', 'Address', 'Status']
        
        if not all(col in df.columns for col in required_columns):
            flash('Invalid template format. Please use the provided template.', 'error')
            return redirect(url_for('lead_client_management'))
        
        success_count = 0
        error_count = 0
        
        for _, row in df.iterrows():
            try:
                existing_client = Client.query.filter_by(email=row['Email']).first()
                if existing_client:
                    error_count += 1
                    continue
                
                new_client = Client(
                    name=row['Name'],
                    email=row['Email'],
                    phone=str(row['Phone']),
                    company=row['Company'],
                    address=row['Address'],
                    status=row['Status']
                )
                db.session.add(new_client)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                continue
        
        db.session.commit()
        flash(f'Successfully added {success_count} clients. {error_count} clients failed.', 'success' if error_count == 0 else 'warning')
        
    except Exception as e:
        flash('Error processing file. Please check the format.', 'error')
    
    return redirect(url_for('lead_client_management'))

@app.route('/download_lead_template')
def download_lead_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Lead Template"
    
    headers = ['Name', 'Email', 'Phone', 'Source', 'Status', 'Notes']
    ws.append(headers)
    
    sample_data = [
        ['John Doe', 'john@example.com', '1234567890', 'Website', 'new', 'Interested in investment plans'],
        ['Jane Smith', 'jane@example.com', '9876543210', 'Referral', 'negotiating', 'Follow up next week']
    ]
    
    for row in sample_data:
        ws.append(row)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='lead_template.xlsx'
    )

@app.route('/download_client_template')
def download_client_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Client Template"
    
    headers = ['Name', 'Email', 'Phone', 'Company', 'Address', 'Status']
    ws.append(headers)
    
    sample_data = [
        ['John Doe Corp', 'contact@johndoe.com', '1234567890', 'JD Corporation', '123 Business St, City', 'active'],
        ['Smith Industries', 'info@smith.com', '9876543210', 'Smith & Co', '456 Industry Ave, Town', 'active']
    ]
    
    for row in sample_data:
        ws.append(row)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='client_template.xlsx'
    )

@app.route('/update-client/<int:client_id>', methods=['POST'])
def update_client(client_id):
    if not session.get('user_id'):
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        client = Client.query.get_or_404(client_id)
        client.name = request.form.get('name')
        client.email = request.form.get('email')
        client.phone = request.form.get('phone')
        client.company = request.form.get('company')
        client.address = request.form.get('address')
        client.status = request.form.get('status')
        assigned_to = request.form.get('assigned_to')
        if assigned_to:
            client.assigned_to = int(assigned_to)
        
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/delete-client/<int:client_id>', methods=['POST'])
def delete_client(client_id):
    if not session.get('user_id'):
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        client = Client.query.get_or_404(client_id)
        db.session.delete(client)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/add-client-service/<int:client_id>', methods=['POST'])
def add_client_service(client_id):
    if not session.get('user_id'):
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        client = Client.query.get_or_404(client_id)
        service = ClientService(
            client_id=client_id,
            name=request.form.get('name'),
            description=request.form.get('description'),
            status=request.form.get('status', 'pending')
        )
        db.session.add(service)
        db.session.commit()
        return jsonify({'success': True, 'service_id': service.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/upload-service-document/<int:service_id>', methods=['POST'])
def upload_service_document(service_id):
    if not session.get('user_id'):
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        service = ClientService.query.get_or_404(service_id)
        if 'document' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'})
        
        file = request.files['document']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'})
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], 'service_documents', filename)
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            file.save(file_path)
            
            document = ServiceDocument(
                service_id=service_id,
                name=filename,
                file_path=os.path.join('service_documents', filename),
                document_type=request.form.get('document_type', 'other')
            )
            db.session.add(document)
            db.session.commit()
            return jsonify({'success': True, 'document_id': document.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get-service-documents/<int:service_id>')
def get_service_documents(service_id):
    if not session.get('user_id'):
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        documents = ServiceDocument.query.filter_by(service_id=service_id).all()
        return jsonify({
            'success': True,
            'documents': [{
                'id': doc.id,
                'name': doc.name,
                'document_type': doc.document_type,
                'upload_date': doc.upload_date.strftime('%Y-%m-%d %H:%M:%S')
            } for doc in documents]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/download-service-document/<int:document_id>')
def download_service_document(document_id):
    if not session.get('user_id'):
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        document = ServiceDocument.query.get_or_404(document_id)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], document.file_path)
        
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'error': 'File not found'}), 404
            
        # Get file extension
        file_ext = os.path.splitext(document.name)[1].lower()
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=document.name,
            mimetype=get_mimetype(file_ext)
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/delete-service-document/<int:document_id>', methods=['POST'])
def delete_service_document(document_id):
    if not session.get('user_id'):
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        document = ServiceDocument.query.get_or_404(document_id)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], document.file_path)
        if os.path.exists(file_path):
            os.remove(file_path)
        db.session.delete(document)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/generate-client-report/<int:client_id>')
def generate_client_report(client_id):
    if not session.get('user_id'):
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        # Get the client
        client = Client.query.get_or_404(client_id)
        
        # Create a PDF buffer
        buffer = BytesIO()
        
        try:
            # Create the PDF document
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )
            
            # Get the styles
            styles = getSampleStyleSheet()
            title_style = styles['Title']
            heading1_style = styles['Heading1']
            heading2_style = styles['Heading2']
            normal_style = styles['Normal']
            
            # Initialize the story list
            story = []
            
            # Add title
            story.append(Paragraph(f"Client Report - {client.name}", title_style))
            story.append(Spacer(1, 20))
            
            # Add client details
            story.append(Paragraph("Client Details", heading1_style))
            story.append(Spacer(1, 12))
            
            # Create client details table
            client_data = [
                ['Name:', client.name],
                ['Email:', client.email or 'N/A'],
                ['Phone:', client.phone or 'N/A'],
                ['Company:', client.company or 'N/A'],
                ['Status:', client.status]
            ]
            
            client_table = Table(client_data, colWidths=[100, 300])
            client_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 6),
            ]))
            
            story.append(client_table)
            story.append(Spacer(1, 20))
            
            # Add services section
            if client.services:
                story.append(Paragraph("Services", heading1_style))
                story.append(Spacer(1, 12))
                
                for service in client.services:
                    story.append(Paragraph(f"Service: {service.name}", heading2_style))
                    
                    # Create service details table
                    service_data = [
                        ['Status:', service.status],
                        ['Description:', service.description or 'N/A']
                    ]
                    
                    service_table = Table(service_data, colWidths=[100, 300])
                    service_table.setStyle(TableStyle([
                        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 0), (-1, -1), 10),
                        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('PADDING', (0, 0), (-1, -1), 6),
                    ]))
                    
                    story.append(service_table)
                    story.append(Spacer(1, 12))
                    
                    # Add documents section with proper error handling
                    if service.documents:
                        story.append(Paragraph("Documents:", heading2_style))
                        doc_data = []
                        for document in service.documents:
                            try:
                                doc_info = f"• {document.name}"
                                if document.document_type:
                                    doc_info += f" ({document.document_type})"
                                if document.upload_date:
                                    doc_info += f" - Uploaded on {document.upload_date.strftime('%Y-%m-%d')}"
                                story.append(Paragraph(doc_info, normal_style))
                            except Exception as doc_error:
                                print(f"Error processing document: {str(doc_error)}")
                                continue
                        story.append(Spacer(1, 12))
                    
                    # Add interactions
                    if service.interactions:
                        story.append(Paragraph("Interactions:", heading2_style))
                        for interaction in service.interactions:
                            try:
                                interaction_text = (
                                    f"• {interaction.interaction_date.strftime('%Y-%m-%d')}: "
                                    f"{interaction.interaction_type}"
                                )
                                story.append(Paragraph(interaction_text, normal_style))
                                if interaction.summary:
                                    story.append(Paragraph(f"  Summary: {interaction.summary}", normal_style))
                            except Exception as int_error:
                                print(f"Error processing interaction: {str(int_error)}")
                                continue
                        story.append(Spacer(1, 12))
            else:
                story.append(Paragraph("No services found for this client.", normal_style))
            
            # Build the PDF
            doc.build(story)
            
            # Get the PDF content
            pdf_content = buffer.getvalue()
            buffer.close()
            
            # Create the response
            response = make_response(pdf_content)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'attachment; filename=client_report_{client.id}.pdf'
            return response
            
        except Exception as pdf_error:
            # Log the specific PDF generation error
            print(f"PDF Generation Error: {str(pdf_error)}")
            if buffer:
                buffer.close()
            return jsonify({'success': False, 'error': f'Error generating PDF: {str(pdf_error)}'}), 500
            
    except Exception as e:
        # Log the general error
        print(f"General Error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/update-service-status/<int:service_id>', methods=['POST'])
def update_service_status(service_id):
    if not session.get('user_id'):
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        service = ClientService.query.get_or_404(service_id)
        status = request.form.get('status')
        
        if status not in ['pending', 'in_progress', 'completed', 'cancelled']:
            return jsonify({'success': False, 'error': 'Invalid status'}), 400
            
        service.status = status
        service.updated_date = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/add-service-interaction/<int:service_id>', methods=['POST'])
def add_service_interaction(service_id):
    if not session.get('user_id'):
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        service = ClientService.query.get_or_404(service_id)
        
        interaction = ServiceInteraction(
            service_id=service_id,
            interaction_type=request.form.get('interaction_type'),
            summary=request.form.get('summary'),
            interaction_date=datetime.strptime(request.form.get('interaction_date'), '%Y-%m-%d'),
            created_by=session['user_id']
        )
        
        if request.form.get('next_followup_date'):
            interaction.next_followup_date = datetime.strptime(request.form.get('next_followup_date'), '%Y-%m-%d')
        
        db.session.add(interaction)
        db.session.commit()
        
        return jsonify({'success': True, 'interaction_id': interaction.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/edit-service-interaction/<int:interaction_id>', methods=['POST'])
def edit_service_interaction(interaction_id):
    if not session.get('user_id'):
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        interaction = ServiceInteraction.query.get_or_404(interaction_id)
        
        interaction.interaction_type = request.form.get('interaction_type')
        interaction.summary = request.form.get('summary')
        interaction.interaction_date = datetime.strptime(request.form.get('interaction_date'), '%Y-%m-%d')
        
        if request.form.get('next_followup_date'):
            interaction.next_followup_date = datetime.strptime(request.form.get('next_followup_date'), '%Y-%m-%d')
        else:
            interaction.next_followup_date = None
        
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/delete-service-interaction/<int:interaction_id>', methods=['POST'])
def delete_service_interaction(interaction_id):
    if not session.get('user_id'):
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        interaction = ServiceInteraction.query.get_or_404(interaction_id)
        db.session.delete(interaction)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get-productivity-data')
def get_productivity_data():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        date_from = request.args.get('from')
        date_to = request.args.get('to')
        
        if not all([date_from, date_to]):
            return jsonify({'success': False, 'error': 'Date range is required'}), 400
        
        # Convert string dates to datetime
        from_date = datetime.strptime(date_from, '%Y-%m-%d')
        to_date = datetime.strptime(date_to, '%Y-%m-%d')
        
        # Initialize data lists
        completion_dates = []
        completion_counts = []
        
        # Get data for each day in range
        current_date = from_date
        while current_date <= to_date:
            date_str = current_date.strftime('%Y-%m-%d')
            completion_dates.append(date_str)
            
            # Count tasks completed on this date
            completed_count = Task.query.filter(
                Task.status == 'completed',
                func.date(Task.created_date) == current_date.date()
            ).count()
            completion_counts.append(completed_count)
            current_date += timedelta(days=1)

        # Collect employee productivity data
        employee_data = []
        employees = Employee.query.all()
        
        if not employees:
            return jsonify({
                'success': True,
                'completion_dates': completion_dates,
                'completion_counts': completion_counts,
                'employee_data': [],
                'message': 'No employees found'
            })

        for employee in employees:
            tasks = Task.query.filter_by(employee_id=employee.id).all()
            
            # Calculate productivity score
            total_tasks = len(tasks)
            completed_tasks = len([t for t in tasks if t.status == 'completed'])
            productivity = round((completed_tasks / total_tasks) * 100, 2) if total_tasks > 0 else 0
            
            # Collect employee's task data
            employee_data.append({
                'employee_name': employee.name,
                'total_tasks': total_tasks,
                'completed_tasks': completed_tasks,
                'productivity': productivity,
                'overdue_tasks': len([t for t in tasks if t.status != 'completed' and t.due_date < datetime.now()])
            })
        
        return jsonify({
            'success': True,
            'completion_dates': completion_dates,
            'completion_counts': completion_counts,
            'employee_data': employee_data
        })
        
    except ValueError as ve:
        # Handle date parsing errors
        return jsonify({
            'success': False, 
            'error': f'Invalid date format: {str(ve)}'
        }), 400
    except Exception as e:
        # Log the full error for server-side debugging
        print(f"Error getting productivity data: {str(e)}")
        return jsonify({
            'success': False, 
            'error': 'An unexpected error occurred while fetching productivity data'
        }), 500

@app.route('/productivity/add-task', methods=['POST'])
def add_productivity_task():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        task = Task(
            title=request.form.get('title'),
            description=request.form.get('description'),
            due_date=datetime.strptime(request.form.get('due_date'), '%Y-%m-%d'),
            priority=request.form.get('priority', 'medium'),
            status=request.form.get('status', 'pending'),
            employee_id=request.form.get('employee_id'),
            assigned_by_id=session['user_id'],
            created_date=datetime.now()
        )
        db.session.add(task)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error adding task: {str(e)}")
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/productivity/update-task/<int:task_id>', methods=['POST'])
def update_productivity_task(task_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        task = Task.query.get_or_404(task_id)
        task.title = request.form.get('title')
        task.description = request.form.get('description')
        task.due_date = datetime.strptime(request.form.get('due_date'), '%Y-%m-%d')
        task.priority = request.form.get('priority')
        task.status = request.form.get('status')
        task.employee_id = request.form.get('employee_id')
        task.updated_date = datetime.now()
        
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error updating task: {str(e)}")
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/productivity/get-task/<int:task_id>')
def get_productivity_task(task_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not authorized'}), 401
    
    try:
        task = Task.query.get_or_404(task_id)
        return jsonify({
            'success': True,
            'id': task.id,
            'title': task.title,
            'description': task.description,
            'due_date': task.due_date.strftime('%Y-%m-%d') if task.due_date else None,
            'priority': task.priority,
            'status': task.status,
            'employee_id': task.employee_id
        })
    except Exception as e:
        print(f"Error getting task: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/add-salary-component', methods=['POST'])
def add_salary_component():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        # Handle both form data and JSON data
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form.to_dict()
        
        # Validate required fields
        required_fields = ['name', 'type', 'is_percentage', 'value', 'employee_id', 'is_taxable']
        for field in required_fields:
            if field not in data:
                return jsonify({'success': False, 'error': f'Missing required field: {field}'}), 400
        
        # Convert and validate values
        try:
            value = float(data['value'])
            employee_id = int(data['employee_id'])
            is_percentage = data['is_percentage'].lower() == 'true'
            is_taxable = data['is_taxable'].lower() == 'true'
        except (ValueError, TypeError) as e:
            return jsonify({'success': False, 'error': f'Invalid value format: {str(e)}'}), 400
        
        # Create new component
        new_component = SalaryComponent(
            name=data['name'],
            type=data['type'],
            is_percentage=is_percentage,
            value=value,
            is_taxable=is_taxable,
            employee_id=employee_id
        )
        
        print("Creating component:", {
            'name': new_component.name,
            'type': new_component.type,
            'is_percentage': new_component.is_percentage,
            'value': new_component.value,
            'is_taxable': new_component.is_taxable,
            'employee_id': new_component.employee_id
        })
        
        db.session.add(new_component)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Salary component added successfully',
            'component': {
                'id': new_component.id,
                'name': new_component.name,
                'type': new_component.type,
                'value': new_component.value,
                'is_percentage': new_component.is_percentage,
                'is_taxable': new_component.is_taxable
            }
        })
        
    except Exception as e:
        print("Error adding salary component:", str(e))  # Debug log
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/get-employee-salary/<int:employee_id>', methods=['GET'])
def get_employee_salary(employee_id):
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        today = datetime.now()
        month = int(request.args.get('month', today.month))
        year = int(request.args.get('year', today.year))
        
        salary = EmployeeSalary.query.filter_by(
            employee_id=employee_id,
            month=month,
            year=year
        ).first()
        
        if not salary:
            return jsonify({'error': 'Salary not found'}), 404
            
        salary_details = []
        for detail in salary.details:
            salary_details.append({
                'id': detail.id,
                'component_name': detail.component.name,
                'amount': detail.amount
            })
            
        return jsonify({
            'id': salary.id,
            'basic_pay': salary.basic_pay,
            'net_salary': salary.net_salary,
            'status': salary.status,
            'details': salary_details
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/generate-payslip/<int:salary_id>', methods=['GET'])
def generate_payslip(salary_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        # Get the salary record
        salary = EmployeeSalary.query.get_or_404(salary_id)
        employee = Employee.query.get_or_404(salary.employee_id)
        
        # Get components for this specific month
        components = SalaryComponent.query.filter_by(employee_id=employee.id).all()
        
        earnings = []
        deductions = []
        total_earnings = salary.basic_pay
        total_deductions = 0
        
        for component in components:
            amount = component.value
            if component.is_percentage:
                amount = (salary.basic_pay * component.value) / 100
                
            if component.type == 'earning':
                earnings.append({
                    'name': component.name,
                    'amount': amount
                })
                total_earnings += amount
            else:
                deductions.append({
                    'name': component.name,
                    'amount': amount
                })
                total_deductions += amount
        
        payslip_data = {
            'employee_name': employee.name,
            'employee_id': employee.id,
            'month': salary.month,
            'year': salary.year,
            'basic_pay': salary.basic_pay,
            'earnings': earnings,
            'deductions': deductions,
            'total_earnings': total_earnings,
            'total_deductions': total_deductions,
            'net_salary': salary.net_salary
        }
        
        return jsonify(payslip_data)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/get-salary-components/<int:employee_id>')
def get_salary_components(employee_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        components = SalaryComponent.query.filter_by(employee_id=employee_id).all()
        return jsonify({
            'success': True,
            'components': [{
                'id': comp.id,
                'name': comp.name,
                'type': comp.type,
                'is_percentage': comp.is_percentage,
                'value': comp.value,
                'is_taxable': comp.is_taxable
            } for comp in components]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/generate-salary/<int:employee_id>', methods=['POST'])
def generate_salary(employee_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        # Handle both form data and JSON data
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form.to_dict()
        
        # Validate required fields
        required_fields = ['basic_pay', 'month', 'year']
        for field in required_fields:
            if field not in data:
                return jsonify({'success': False, 'error': f'Missing required field: {field}'}), 400
        
        basic_pay = float(data['basic_pay'])
        month = int(data['month'])
        year = int(data['year'])
        
        # Get or create salary record
        salary = EmployeeSalary.query.filter_by(
            employee_id=employee_id,
            month=month,
            year=year
        ).first()
        
        if not salary:
            salary = EmployeeSalary(
                employee_id=employee_id,
                month=month,
                year=year,
                basic_pay=basic_pay,
                status='pending'
            )
            db.session.add(salary)
        else:
            salary.basic_pay = basic_pay
            
        # Calculate net salary based on components
        components = SalaryComponent.query.filter_by(employee_id=employee_id).all()
        net_salary = basic_pay
        
        # Delete existing salary details
        SalaryDetail.query.filter_by(salary_id=salary.id).delete()
        
        # Calculate and add new salary details
        for component in components:
            amount = component.value
            if component.is_percentage:
                amount = (basic_pay * component.value) / 100
                
            if component.type == 'deduction':
                net_salary -= amount
            else:
                net_salary += amount
                
            detail = SalaryDetail(
                salary_id=salary.id,
                component_id=component.id,
                amount=amount
            )
            db.session.add(detail)
            
        salary.net_salary = net_salary
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Salary generated successfully',
            'salary': {
                'id': salary.id,
                'basic_pay': salary.basic_pay,
                'net_salary': salary.net_salary,
                'status': salary.status
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/process-all-salaries', methods=['POST'])
def process_all_salaries():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        # Handle both form data and JSON data
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form.to_dict()
        
        # Validate required fields
        required_fields = ['month', 'year']
        for field in required_fields:
            if field not in data:
                return jsonify({'success': False, 'error': f'Missing required field: {field}'}), 400
        
        month = int(data['month'])
        year = int(data['year'])
        
        # Get all pending salaries for the specified month/year
        pending_salaries = EmployeeSalary.query.filter_by(
            status='pending',
            month=month,
            year=year
        ).all()
        
        if not pending_salaries:
            return jsonify({'success': False, 'error': 'No pending salaries found'}), 400
        
        # Process each salary
        processed_count = 0
        for salary in pending_salaries:
            salary.status = 'processed'
            salary.processed_date = datetime.utcnow()
            salary.processed_by = session['user_id']
            processed_count += 1
        
        print(f"Processed {processed_count} salaries")  # Debug log
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully processed {processed_count} salaries',
            'processed_count': processed_count
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/payslip/<int:salary_id>')
def view_payslip(salary_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    try:
        # Get the salary record
        salary = EmployeeSalary.query.get_or_404(salary_id)
        employee = Employee.query.get_or_404(salary.employee_id)
        
        # Get components for this specific month
        components = SalaryComponent.query.filter_by(employee_id=employee.id).all()
        
        earnings = []
        deductions = []
        total_earnings = salary.basic_pay
        total_deductions = 0
        
        for component in components:
            amount = component.value
            if component.is_percentage:
                amount = (salary.basic_pay * component.value) / 100
                
            if component.type == 'earning':
                earnings.append({
                    'name': component.name,
                    'amount': amount
                })
                total_earnings += amount
            else:
                deductions.append({
                    'name': component.name,
                    'amount': amount
                })
                total_deductions += amount
        
        salary_date = datetime(salary.year, salary.month, 1)
        net_salary = total_earnings - total_deductions
        
        # Add num2words to template context
        app.jinja_env.globals.update(num2words=num2words)
        
        # Render HTML template
        html = render_template('payslip_pdf.html',
                            employee=employee,
                            salary=salary,
                            earnings=earnings,
                            deductions=deductions,
                            total_earnings=total_earnings,
                            total_deductions=total_deductions,
                            net_salary=net_salary,
                            salary_date=salary_date,
                            generation_date=datetime.now())
        
        # Configure pdfkit options
        options = {
            'page-size': 'A4',
            'margin-top': '0.75in',
            'margin-right': '0.75in',
            'margin-bottom': '0.75in',
            'margin-left': '0.75in',
            'encoding': 'UTF-8',
            'enable-local-file-access': None
        }
        
        try:
            config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
            pdf = pdfkit.from_string(html, False, options=options, configuration=config)
            
            response = make_response(pdf)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'inline; filename=payslip_{employee.name}_{salary_date.strftime("%B_%Y")}.pdf'
            
            return response
        except Exception as e:
            print(f"PDF generation failed: {str(e)}")
            return html
            
    except Exception as e:
        print(f"Error: {str(e)}")
        flash(f'Error generating payslip: {str(e)}', 'error')
        return redirect(url_for('payslip'))

@app.route('/get-attendance')
def get_attendance():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    employee_id = request.args.get('employee_id', type=int)
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)
    
    if not all([employee_id, year, month]):
        return jsonify({'success': False, 'error': 'Missing parameters'})
    
    try:
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1) - timedelta(days=1)
        
        attendance_records = Attendance.query.filter(
            Attendance.employee_id == employee_id,
            Attendance.date >= start_date,
            Attendance.date <= end_date
        ).all()
        
        calendar_days = []
        current_date = start_date
        while current_date <= end_date:
            attendance = next(
                (a for a in attendance_records if a.date == current_date.date()),
                None
            )
            calendar_days.append({
                'date': current_date,
                'status': attendance.status if attendance else 'unmarked'
            })
            current_date += timedelta(days=1)
        
        return jsonify({
            'success': True,
            'month': start_date.strftime('%B %Y'),
            'attendance': calendar_days
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/add-attendance/<int:employee_id>', methods=['POST'])
def add_attendance(employee_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    try:
        date_str = request.form.get('date')
        status = request.form.get('status')
        comments = request.form.get('comments', '')
        
        if not all([date_str, status]):
            return jsonify({'success': False, 'error': 'Missing required fields'})
        
        # Try parsing the date in different formats
        try:
            attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            try:
                # Parse RFC format date
                attendance_date = datetime.strptime(date_str, '%a, %d %b %Y %H:%M:%S GMT').date()
            except ValueError:
                return jsonify({'success': False, 'error': f'Invalid date format: {date_str}'})
        
        existing_attendance = Attendance.query.filter_by(
            employee_id=employee_id,
            date=attendance_date
        ).first()
        
        if existing_attendance:
            existing_attendance.status = status
            existing_attendance.comments = comments
        else:
            new_attendance = Attendance(
                employee_id=employee_id,
                date=attendance_date,
                status=status,
                comments=comments
            )
            db.session.add(new_attendance)
        
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

# Analytics Helper Functions
def get_main_chart_data(module, start_date, end_date):
    try:
        if module == 'recruitment':
            # Get job applications data
            applications = Job_Application.query.filter(
                Job_Application.created_date.between(start_date, end_date)
            ).all()
            
            dates = []
            counts = []
            current_date = start_date
            while current_date <= end_date:
                date_str = current_date.strftime('%Y-%m-%d')
                count = sum(1 for app in applications if app.created_date.date() == current_date.date())
                dates.append(date_str)
                counts.append(count)
                current_date += timedelta(days=1)
            
            return {
                'labels': dates,
                'datasets': [{
                    'label': 'Job Applications',
                    'data': counts,
                    'backgroundColor': 'rgba(54, 162, 235, 0.5)',
                    'borderColor': 'rgba(54, 162, 235, 1)',
                    'borderWidth': 1
                }]
            }
            
        elif module == 'employee':
            # Get task completion data by employee
            employees = Employee.query.all()
            tasks = Task.query.filter(
                Task.created_date.between(start_date, end_date)
            ).all()
            
            employee_names = []
            completion_rates = []
            
            for emp in employees:
                emp_tasks = [t for t in tasks if t.employee_id == emp.id]
                if emp_tasks:
                    completed = sum(1 for t in emp_tasks if t.status == 'completed')
                    rate = (completed / len(emp_tasks)) * 100
                    employee_names.append(emp.name)
                    completion_rates.append(round(rate, 1))
            
            return {
                'labels': employee_names,
                'datasets': [{
                    'label': 'Task Completion Rate (%)',
                    'data': completion_rates,
                    'backgroundColor': 'rgba(75, 192, 192, 0.5)',
                    'borderColor': 'rgba(75, 192, 192, 1)',
                    'borderWidth': 1
                }]
            }
            
        elif module == 'budget':
            # Get expense data by category
            expenses = Expense.query.filter(
                Expense.date.between(start_date, end_date)
            ).all()
            
            categories = {}
            for expense in expenses:
                categories[expense.category] = categories.get(expense.category, 0) + expense.amount
            
            return {
                'labels': list(categories.keys()),
                'datasets': [{
                    'label': 'Expenses by Category',
                    'data': list(categories.values()),
                    'backgroundColor': [
                        'rgba(255, 99, 132, 0.5)',
                        'rgba(54, 162, 235, 0.5)',
                        'rgba(255, 206, 86, 0.5)',
                        'rgba(75, 192, 192, 0.5)'
                    ]
                }]
            }
            
        elif module == 'social':
            # Get social media metrics
            metrics = {
                'LinkedIn': random.randint(50, 100),
                'Twitter': random.randint(30, 80),
                'Facebook': random.randint(40, 90),
                'Instagram': random.randint(60, 120)
            }
            
            return {
                'labels': list(metrics.keys()),
                'datasets': [{
                    'label': 'Social Media Engagement',
                    'data': list(metrics.values()),
                    'backgroundColor': 'rgba(153, 102, 255, 0.5)',
                    'borderColor': 'rgba(153, 102, 255, 1)',
                    'borderWidth': 1
                }]
            }
        
        return {
            'labels': [],
            'datasets': [{
                'label': 'No Data',
                'data': [],
                'backgroundColor': 'rgba(75, 192, 192, 0.5)',
                'borderColor': 'rgba(75, 192, 192, 1)',
                'borderWidth': 1
            }]
        }
    except Exception as e:
        print(f"Error in get_main_chart_data: {str(e)}")
        return {
            'labels': [],
            'datasets': [{
                'label': 'Error',
                'data': [],
                'backgroundColor': 'rgba(255, 99, 132, 0.5)'
            }]
        }

def get_trend_chart_data(module, start_date, end_date):
    try:
        dates = []
        current_date = start_date
        while current_date <= end_date:
            dates.append(current_date.strftime('%Y-%m-%d'))
            current_date += timedelta(days=1)
        
        if module == 'recruitment':
            # Get daily application trends
            applications = Job_Application.query.filter(
                Job_Application.created_date.between(start_date, end_date)
            ).all()
            
            daily_counts = []
            for date in dates:
                date_obj = datetime.strptime(date, '%Y-%m-%d')
                count = sum(1 for app in applications if app.created_date.date() == date_obj.date())
                daily_counts.append(count)
            
            return {
                'labels': dates,
                'datasets': [{
                    'label': 'Daily Applications',
                    'data': daily_counts,
                    'fill': False,
                    'borderColor': 'rgb(75, 192, 192)',
                    'tension': 0.1
                }]
            }
            
        elif module == 'employee':
            # Get daily task completion trend
            tasks = Task.query.filter(
                Task.created_date.between(start_date, end_date)
            ).all()
            
            completion_rates = []
            for date in dates:
                date_obj = datetime.strptime(date, '%Y-%m-%d')
                day_tasks = [t for t in tasks if t.created_date.date() == date_obj.date()]
                if day_tasks:
                    completed = sum(1 for t in day_tasks if t.status == 'completed')
                    rate = (completed / len(day_tasks)) * 100
                    completion_rates.append(round(rate, 1))
                else:
                    completion_rates.append(0)
            
            return {
                'labels': dates,
                'datasets': [{
                    'label': 'Daily Task Completion Rate (%)',
                    'data': completion_rates,
                    'fill': False,
                    'borderColor': 'rgb(54, 162, 235)',
                    'tension': 0.1
                }]
            }
        
        return {
            'labels': dates,
            'datasets': [{
                'label': 'No Trend Data',
                'data': [0] * len(dates),
                'fill': False,
                'borderColor': 'rgb(75, 192, 192)',
                'tension': 0.1
            }]
        }
    except Exception as e:
        print(f"Error in get_trend_chart_data: {str(e)}")
        return {
            'labels': [],
            'datasets': [{
                'label': 'Error',
                'data': [],
                'fill': False,
                'borderColor': 'rgb(255, 99, 132)'
            }]
        }

def get_distribution_chart_data(module, start_date, end_date):
    try:
        if module == 'recruitment':
            # Get application status distribution
            applications = Job_Application.query.filter(
                Job_Application.created_date.between(start_date, end_date)
            ).all()
            
            status_counts = {}
            for app in applications:
                status_counts[app.status] = status_counts.get(app.status, 0) + 1
            
            return {
                'labels': list(status_counts.keys()),
                'datasets': [{
                    'data': list(status_counts.values()),
                    'backgroundColor': [
                        'rgba(255, 206, 86, 0.5)',
                        'rgba(75, 192, 192, 0.5)',
                        'rgba(255, 99, 132, 0.5)',
                        'rgba(54, 162, 235, 0.5)'
                    ]
                }]
            }
            
        elif module == 'employee':
            # Get task status distribution
            tasks = Task.query.filter(
                Task.created_date.between(start_date, end_date)
            ).all()
            
            status_counts = {}
            for task in tasks:
                status_counts[task.status] = status_counts.get(task.status, 0) + 1
            
            return {
                'labels': list(status_counts.keys()),
                'datasets': [{
                    'data': list(status_counts.values()),
                    'backgroundColor': [
                        'rgba(75, 192, 192, 0.5)',
                        'rgba(255, 99, 132, 0.5)',
                        'rgba(255, 206, 86, 0.5)'
                    ]
                }]
            }
        
        return {
            'labels': [],
            'datasets': [{
                'data': [],
                'backgroundColor': []
            }]
        }
    except Exception as e:
        print(f"Error in get_distribution_chart_data: {str(e)}")
        return {
            'labels': ['Error'],
            'datasets': [{
                'data': [0],
                'backgroundColor': ['rgba(255, 99, 132, 0.5)']
            }]
        }

def get_summary_data(module, start_date, end_date):
    try:
        if module == 'recruitment':
            # Get recruitment summary
            applications = Job_Application.query.filter(
                Job_Application.created_date.between(start_date, end_date)
            ).all()
            
            total = len(applications)
            if total > 0:
                status_counts = {}
                for app in applications:
                    status_counts[app.status] = status_counts.get(app.status, 0) + 1
                
                return {
                    'Total Applications': total,
                    'Pending Review': status_counts.get('pending', 0),
                    'Shortlisted': status_counts.get('shortlisted', 0),
                    'Rejected': status_counts.get('rejected', 0),
                    'Success Rate': f"{(status_counts.get('shortlisted', 0) / total * 100):.1f}%"
                }
            return {
                'Total Applications': 0,
                'Pending Review': 0,
                'Shortlisted': 0,
                'Rejected': 0,
                'Success Rate': '0%'
            }
            
        elif module == 'employee':
            # Get employee performance summary
            tasks = Task.query.filter(
                Task.created_date.between(start_date, end_date)
            ).all()
            
            total = len(tasks)
            if total > 0:
                completed = sum(1 for t in tasks if t.status == 'completed')
                pending = sum(1 for t in tasks if t.status == 'pending')
                overdue = sum(1 for t in tasks if t.status != 'completed' and t.due_date and t.due_date < datetime.now())
                
                return {
                    'Total Tasks': total,
                    'Completed': completed,
                    'Pending': pending,
                    'Overdue': overdue,
                    'Completion Rate': f"{(completed / total * 100):.1f}%"
                }
            return {
                'Total Tasks': 0,
                'Completed': 0,
                'Pending': 0,
                'Overdue': 0,
                'Completion Rate': '0%'
            }
            
        elif module == 'budget':
            # Get budget summary
            expenses = Expense.query.filter(
                Expense.date.between(start_date, end_date)
            ).all()
            
            total_expense = sum(expense.amount for expense in expenses)
            categories = {}
            for expense in expenses:
                categories[expense.category] = categories.get(expense.category, 0) + expense.amount
            
            highest_category = max(categories.items(), key=lambda x: x[1])[0] if categories else 'None'
            
            return {
                'Total Expenses': f"${total_expense:,.2f}",
                'Number of Transactions': len(expenses),
                'Highest Expense Category': highest_category,
                'Average Transaction': f"${(total_expense/len(expenses) if expenses else 0):,.2f}"
            }
        
        return {'No Data Available': 0}
    except Exception as e:
        print(f"Error in get_summary_data: {str(e)}")
        return {'Error': str(e)}

if __name__ == '__main__':
    with app.app_context():
        # Only create tables if they don't exist
        db.create_all()
        
        # Create default admin user if it doesn't exist
        admin_user = User.query.filter_by(email='admin@example.com').first()
        if not admin_user:
            admin_user = User(
                email='admin@example.com',
                password=generate_password_hash('admin123'),
                name='Admin',
                is_manager=True,
                hierarchy_level=3  # Admin level
            )
            db.session.add(admin_user)
        
        # Create default document categories if they don't exist
        default_categories = [
            {'name': 'HR Documents', 'is_confidential': True},
            {'name': 'Financial Reports', 'is_confidential': True},
            {'name': 'Employee Records', 'is_confidential': True},
            {'name': 'Company Policies', 'is_confidential': False},
            {'name': 'Training Materials', 'is_confidential': False},
            {'name': 'General Documents', 'is_confidential': False}
        ]
        
        for category_data in default_categories:
            existing_category = DocumentCategory.query.filter_by(name=category_data['name']).first()
            if not existing_category:
                category = DocumentCategory(
                    name=category_data['name'],
                    is_confidential=category_data['is_confidential'],
                    access_code='1234' if category_data['is_confidential'] else None
                )
                db.session.add(category)
        
        # Create default leave types if they don't exist
        default_leave_types = [
            {'name': 'Annual Leave', 'default_days': 20, 'color_code': '#28a745'},
            {'name': 'Sick Leave', 'default_days': 12, 'color_code': '#dc3545'},
            {'name': 'Casual Leave', 'default_days': 6, 'color_code': '#ffc107'},
            {'name': 'Maternity Leave', 'default_days': 90, 'color_code': '#e83e8c'},
            {'name': 'Paternity Leave', 'default_days': 5, 'color_code': '#6f42c1'},
            {'name': 'Unpaid Leave', 'default_days': 0, 'color_code': '#6c757d'}
        ]
        
        for leave_type_data in default_leave_types:
            existing_leave_type = LeaveType.query.filter_by(name=leave_type_data['name']).first()
            if not existing_leave_type:
                leave_type = LeaveType(
                    name=leave_type_data['name'],
                    default_days=leave_type_data['default_days'],
                    color_code=leave_type_data['color_code']
                )
                db.session.add(leave_type)
        
        try:
            db.session.commit()
        except Exception as e:
            print(f"Error creating default data: {e}")
            db.session.rollback()
            
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 8080)))